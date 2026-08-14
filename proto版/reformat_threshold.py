#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
一站式比对：拉取实时数据 → 阈值口径比对
一次运行输出阈值口径 xlsx 报告（带时间戳，永不覆盖）
需要严格相等时把 compare_config.yaml thresholds 调成 0 即可
"""
import json, csv, os, datetime, hashlib, hmac, subprocess, time
import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
from fetch_cn_pb import fetch_cn_pb, normalize_cn, normalize_in   # 国内 proto detail + pb->旧结构 + 海外补ts

MATCH_TOLERANCE = datetime.timedelta(minutes=10)

# ====== 接口鉴权配置 ======
PASSWORD_CN = '49ff9a4e5e8bd5e8ce9e057c5adc5d2d'
PASSWORD_IN = '923ffbda8b65bf0f8e126824d050887a'
TOKEN_CN = 'cc920d85f8fbb762b6c705375add6c32'
TOKEN_IN = 'b88b7a5375e293671270016fe556a4b5'

# ====== 路径配置 ======
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.join(SCRIPT_DIR, '..', 'data')
CITY_CSV = os.path.join(BASE, '天气一致性测试城市_热门城市筛选.csv')
OUT_DIR = os.path.join(BASE, '比对结果')
TIMESTAMP = datetime.datetime.now().strftime('%Y%m%d_%H%M')
CITY_RANK = {}  # 城市序号，偏差相同时偏远地区优先
XLSX_THRESHOLD = os.path.join(OUT_DIR, f'一致性比对报告_阈值口径_{TIMESTAMP}.xlsx')
CONFIG_PATH = os.environ.get('CONFIG_PATH') or os.path.join(SCRIPT_DIR, 'compare_config.yaml')

# ====== 读取阈值配置 ======
config = yaml.safe_load(open(CONFIG_PATH, encoding='utf-8'))
THRESHOLDS = config['thresholds']
WIND_CFG = config['wind_convert']
MODULES = config['modules']
WEATHER_MAP = config.get('weather_mapping', {})
WTH_TEXTS = WEATHER_MAP.get('texts', {})

# ====== 降水量等级配置 ======
_rain_raw = config.get('rain_thresholds', {})
RAIN_TH = []          # [(上界, 等级名), ...]
RAIN_NAMES = []       # [等级名, ...]
for name, th in _rain_raw.items():
    if th == '~' or th is None:
        RAIN_TH.append((float('inf'), name))
    else:
        RAIN_TH.append((float(th), name))
    RAIN_NAMES.append(name)
# 按上界升序排序
RAIN_TH.sort(key=lambda x: x[0])

# ====== 数据清洗配置（物理合理范围）======
_CLEAN_CFG = config.get('data_cleaning', {}) or {}
CLEAN_ENABLED = _CLEAN_CFG.get('enabled', False)
CLEAN_RANGES = _CLEAN_CFG.get('ranges', {}) or {}

# ====== 24小时时效分段 ======
# 优先读配置 compare_config.yaml -> hourly_segments; 无则用默认三段
# lead=(predictTime-updatetime)/3600, 命中 (min_hour, max_hour] 归入该段
_seg_cfg = config.get('hourly_segments') or []
if _seg_cfg:
    PERIODS_24H = [(s.get('label'), s.get('min_hour'), s.get('max_hour')) for s in _seg_cfg]
else:
    PERIODS_24H = [
        ('短时效(1-6h)', 0, 6),
        ('中时效(7-12h)', 6, 12),
        ('长时效(13-24h)', 12, 24),
    ]

# =========================================================
# 第一部分：API 请求
# =========================================================

def cn_key(lat, lon):
    return hashlib.md5((PASSWORD_CN + '0' + lat + lon).encode()).hexdigest()

def in_key(lat, lon):
    return hmac.new(PASSWORD_IN.encode(), ('0' + lat + lon).encode(), hashlib.sha256).hexdigest()

def cn_url(lat, lon):
    return f'http://coapi.moji.com/whapi/v2/weather?timestamp=0&language=zh-CN&token={TOKEN_CN}&lat={lat}&lon={lon}&key={cn_key(lat, lon)}'

def in_url_full(lat, lon):
    return f'https://datasw1.api.moweather.com/whapi/in/weather?token={TOKEN_IN}&lon={lon}&lat={lat}&lang=zh-CN&current=1&hourly=24&daily=15&aqi=1&metric=true&ts=0&key={in_key(lat, lon)}'

def fetch(url, retry=2):
    """请求接口，重试 retry 次"""
    for _ in range(retry + 1):
        r = subprocess.run(['curl', '-sk', '--max-time', '25', url], capture_output=True, text=True)
        try:
            d = json.loads(r.stdout)
            if d.get('code') == 0:
                return d.get('data', {})
        except:
            pass
        time.sleep(0.5)
    return None

def fetch_city(name, lon, lat):
    """拉单个城市的国内+国际数据，任一失败对应项为 None。供一次性脚本和定时脚本复用
    国内: proto detail 接口(POST, pb 二进制), fetch_cn_pb 返回 pb 完整 dict(含 detail, 存原始用)
    海外: moweather json 接口, 不变"""
    cn_data = fetch_cn_pb(lon, lat)
    in_data = fetch(in_url_full(lat, lon))
    return cn_data, in_data

# =========================================================
# 第二部分：比对工具函数
# =========================================================

def num(v):
    if v is None: return None
    try: return float(v)
    except: return None

def wind_convert(v):
    if v is None or not WIND_CFG.get('enabled'): return v
    return round(v / WIND_CFG['divisor'], 2)

def text_diff(cv, iv):
    if cv is None or iv is None: return ''
    return '一致' if cv == iv else '不一致'

def rain_level(v):
    """降水量mm → 等级索引(从0开始), 等级定义来自 compare_config.yaml rain_thresholds"""
    if v is None: return None
    for i, (th, _) in enumerate(RAIN_TH):
        if v < th:
            return i
    return len(RAIN_TH) - 1

def get_threshold(field):
    """按字段名匹配阈值（支持'温度(最高)'等带括号字段）"""
    for k, v in THRESHOLDS.items():
        if k in field: return v
    return None

def get_range(field):
    """按字段名匹配清洗范围(支持'温度(最高)'等带括号字段), 返回 (键名, 范围dict) 或 None
    匹配规则同 get_threshold: 键名出现在字段名里即生效"""
    for k, v in CLEAN_RANGES.items():
        if k in field:
            return k, v
    return None

def clean_value(field, cv, iv_conv):
    """数据清洗: 值超出物理合理范围视为脏数据
    返回 (是否剔除, 备注原因); 未配置范围或值在范围内返回 (False, '')"""
    rng = get_range(field)
    if rng is None:
        return False, ''
    kname, r = rng
    lo = r.get('min'); hi = r.get('max'); unit = r.get('unit', '')
    if lo is None or hi is None:
        return False, ''
    bad = []
    if cv is not None and (cv < lo or cv > hi):
        bad.append(f'国内{cv}{unit}')
    if iv_conv is not None and (iv_conv < lo or iv_conv > hi):
        bad.append(f'海外{iv_conv}{unit}')
    if bad:
        return True, f'清洗剔除: 超{kname}范围{lo}~{hi}{unit}({", ".join(bad)})'
    return False, ''

def _parse_utc(s):
    """UTC时间字符串 -> datetime(UTC)。支持 %Y-%m-%d %H:%M:%S 和 %Y-%m-%d"""
    if not s:
        return None
    try:
        if ' ' in s:
            return datetime.datetime.strptime(s, '%Y-%m-%d %H:%M:%S').replace(tzinfo=datetime.timezone.utc)
        else:
            return datetime.datetime.strptime(s, '%Y-%m-%d').replace(tzinfo=datetime.timezone.utc)
    except ValueError:
        return None

def _match_nearest(items, target_utc_str, tolerance=MATCH_TOLERANCE):
    """在国际列表中找 _utc 与 target_utc_str 最接近且 ≤ tolerance 的条目。
    返回 (匹配条目, 时间差秒数); 匹配失败返回 (None, None)"""
    target_dt = _parse_utc(target_utc_str)
    if target_dt is None:
        return None, None
    best = None
    best_diff = None
    for item in items:
        item_dt = _parse_utc(item.get('_utc'))
        if item_dt is None:
            continue
        diff = abs((target_dt - item_dt).total_seconds())
        if diff <= tolerance.total_seconds():
            if best is None or diff < best_diff:
                best = item
                best_diff = diff
    return best, best_diff

def get_period_label(source, idx, base_ts=None, predict_ts=None):
    """24小时时效分段: 优先按真实预报时效 (base_ts=updatetime, predict_ts=predictTime, 单位秒UTC),
    lead=(predict_ts - base_ts)/3600; 无时间戳时回退到按下标分段。
    边界: lead > min_hour AND lead <= max_hour (与参考版一致, 左开右闭)"""
    if source != 'hourly':
        return ''
    if base_ts is not None and predict_ts is not None and base_ts > 0 and predict_ts > 0:
        lead = (predict_ts - base_ts) / 3600
        for label, start, end in PERIODS_24H:
            if start < lead <= end:
                return label
        return ''
    # 回退: 按下标分段(左闭右开, 现有行为)
    for label, start, end in PERIODS_24H:
        if start <= idx < end:
            return label
    return ''

def calc_weather_deviation(cn_text, intl_text):
    """
    天气现象语义映射偏差计算（五分制评分 → 偏差值）
    将中文天气文字映射到(大类,量级,高影响标记)，按评分规则打分

    评分规则（5=最佳 → 0=最差，仅5分算一致）：
      5分  主天气一致+量级一致                     → 完全匹配
      4分  主天气一致+量级差1级                     → 轻微量级偏差
      3分  主天气不一致（晴↔多云等），均非高影响      → 主天气不一致
      2分  主天气错判（量级差≥2），或降水vs非降水    → 明显偏差
      1分  涉及高影响天气+同大类                   → 高影响偏差
      0分  高影响天气+不同大类                     → 高影响漏报/错判

    返回 (偏差, '一致'/'不一致')
    """
    if cn_text is None or intl_text is None:
        return (None, '')

    a = WTH_TEXTS.get(cn_text)
    b = WTH_TEXTS.get(intl_text)
    ok_min = WEATHER_MAP.get('ok_min_score', 5)

    # 未识别的天气文字 → 按缺数据处理，不计入一致率分母（与参考版一致）
    if a is None or b is None:
        return (None, '')

    level_diff = abs(a['level'] - b['level'])

    if a['cat'] == b['cat']:
        if level_diff == 0:
            score = WEATHER_MAP['score_same_cat_same_level']
        elif level_diff == 1:
            score = WEATHER_MAP['score_same_cat_level_diff_1']
        else:
            # level_diff >= 2
            if a['hi'] or b['hi']:
                score = WEATHER_MAP['score_hi_same_cat']
            else:
                score = WEATHER_MAP['score_same_cat_level_diff_ge2']
    else:
        if a['hi'] or b['hi']:
            score = WEATHER_MAP['score_hi_diff_cat']
        else:
            # 一方雨/雪一方非降水 → 比纯云量差异更严重
            is_precip = lambda t: str(t.get('cat', '')).startswith('code_8') or str(t.get('cat', '')).startswith('code_16')
            if is_precip(a) != is_precip(b):
                score = WEATHER_MAP.get('score_diff_cat_precip', 2)
            else:
                score = WEATHER_MAP['score_diff_cat_no_hi']

    deviation = ok_min - score
    ok = '一致' if score >= ok_min else '不一致'
    return (deviation, ok)

def cmp_point(city, module, field, ts, cnv, iv, spec, period=''):
    """
    单个数据点比对: 阈值判断（|diff|<=threshold）
    返回 10 元组
    """
    typ = spec.get('type', 'numeric'); note = spec.get('note', '')
    if typ == 'wind':
        cv = num(cnv)
        iv_conv = wind_convert(num(iv))  # 国际风速总是 km/h, 需换算
        # 国内24小时风速存的是 km/h 原值, 也需换算 (实况/15天存的是 m/s, 不需)
        if cv is not None and module == '24小时':
            cv = wind_convert(cv)
        if WIND_CFG.get('enabled') and iv is not None:
            note = f"海外原{iv} ÷{WIND_CFG['divisor']}"
    elif typ == 'weather':
        cv = cnv; iv_conv = iv
    else:
        cv = num(cnv); iv_conv = num(iv)

    if typ == 'weather':
        diff, ok = calc_weather_deviation(cnv, iv)
        return (city, module, field, ts, cnv, iv, diff, ok, '按语义映射比对', period)

    if typ == 'rain_level':
        cv = num(cnv); iv_conv = num(iv)
        if cv is None or iv_conv is None:
            return (city, module, field, ts, cnv, iv, '', '缺数据', note, period)
        # ---- 数据清洗: 超物理范围的脏数据剔除, 不计入一致率分母 ----
        if CLEAN_ENABLED:
            bad, why = clean_value(field, cv, iv_conv)
            if bad:
                return (city, module, field, ts, cnv, iv, '', '清洗剔除', why, period)
        cl = rain_level(cv); il = rain_level(iv_conv)
        diff = round(abs(cv - iv_conv), 2)
        ok = '一致' if cl == il else '不一致'
        cn_lv = RAIN_NAMES[cl] if cl is not None else '?'
        in_lv = RAIN_NAMES[il] if il is not None else '?'
        return (city, module, field, ts, cnv, iv, diff, ok, f'国内{cn_lv} vs 海外{in_lv}', period)

    if cv is None or iv_conv is None:
        return (city, module, field, ts, cnv, iv, '', '缺数据', note, period)

    # ---- 数据清洗: 超物理范围的脏数据剔除, 不计入一致率分母 ----
    if CLEAN_ENABLED:
        bad, why = clean_value(field, cv, iv_conv)
        if bad:
            return (city, module, field, ts, cnv, iv, '', '清洗剔除', why, period)

    diff = round(abs(cv - iv_conv), 2)
    th = get_threshold(field)
    ok = '一致' if (th is not None and abs(diff) <= th) else '不一致'
    return (city, module, field, ts, cv, iv_conv, diff, ok, note, period)

def build_points(city, cn, ind):
    """遍历所有模块/字段，返回该城市比对数据点列表
    cn 若是国内 proto detail 的 pb 完整 dict(含 detail key), 先 normalize 成旧结构再比对"""
    if cn and isinstance(cn, dict) and 'detail' in cn:
        cn = normalize_cn(cn)
    if ind:
        # P0修复: 必须传国内时区让海外 predict_time/predict_date 转当地时间,
        # 否则海外补的是UTC, 与国内(当地)精确字符串匹配错位8小时(北京18:00配UTC18:00)。
        # normalize_cn 的 _meta.timezone 是该城市时区偏移(如北京8)。与 CSV 工作流(convert_raw_to_csv)同口径。
        tz_hours = cn.get('_meta', {}).get('timezone') if isinstance(cn, dict) else None
        ind = normalize_in(ind, tz_hours=tz_hours)
    P = []
    for module, mspec in MODULES.items():
        source = mspec['source']; fields = mspec['fields']
        if mspec.get('multi'):
            cn_arr = cn.get(source, [])
            ind_arr = ind.get(source, [])   # 不预截断, 保证 nearest 在全量国际数组中匹配
            ts_key = mspec.get('ts_key'); lim = mspec.get('limit', 99)

            if source == 'hourly':
                # 24小时: UTC最近匹配 + 10分钟容差, 按真实预报时效分段 (与参考版一致)
                base_ts_val = cn.get('_meta', {}).get('updatetime')
                matched = []
                for a in cn_arr:
                    b, _ = _match_nearest(ind_arr, a.get('_utc'))
                    if b is None:
                        continue
                    predict_ts = a.get('_predict_ts')
                    matched.append((a.get(ts_key), a, b, predict_ts, base_ts_val))
                    if len(matched) >= lim:
                        break
                for k, (ts, a, b, predict_ts, base_ts_v) in enumerate(matched):
                    period = get_period_label(source, k, base_ts_v, predict_ts)
                    if not period:
                        continue   # 超出0-24h时效范围的记录不纳入小时预报统计(与参考版一致)
                    for field, spec in fields.items():
                        if spec.get('compare') is False:
                            continue   # 只拉取不比对(紫外线/能见度/风向)
                        P.append(cmp_point(city, module, field, ts, a.get(spec['cn']), b.get(spec['intl']), spec, period))
            else:
                # 15天: 按日期字符串匹配(同现有逻辑)
                ind_map = {}
                for b in ind_arr[:lim]:
                    t = b.get(ts_key)
                    if t is not None:
                        ind_map[t] = b
                matched = []
                for a in cn_arr:
                    t = a.get(ts_key)
                    if t is not None and t in ind_map:
                        matched.append((t, a, ind_map[t]))
                    if len(matched) >= lim:
                        break
                for k, (ts, a, b) in enumerate(matched):
                    period = get_period_label(source, k)
                    for field, spec in fields.items():
                        if spec.get('compare') is False:
                            continue   # 只拉取不比对(紫外线/能见度/风向)
                        P.append(cmp_point(city, module, field, ts, a.get(spec['cn']), b.get(spec['intl']), spec, period))
        else:
            cn_mod = cn.get(source, {}); ind_mod = ind.get(source, {})
            ts = mspec.get('ts_label', '')
            for field, spec in fields.items():
                if spec.get('compare') is False:
                    continue   # 只拉取不比对(紫外线/能见度/风向)
                P.append(cmp_point(city, module, field, ts, cn_mod.get(spec['cn']), ind_mod.get(spec['intl']), spec, ''))
    return P

def write_detail_csv(points, out_dir, name_prefix, batch_time=None):
    """按模块分文件写数据明细CSV(每个模块一个文件), 每行带写入时间(批次标识)
    points: cmp_point 返回的 10 元组列表
    out_dir: 输出目录
    name_prefix: 文件名前缀, 拼成 {name_prefix}_{模块}.csv
    batch_time: 批次写入时间(同批所有行一致, 便于追溯批次), 默认当前时刻
    返回写入的文件路径列表"""
    if batch_time is None:
        batch_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    by_mod = {}
    for p in points:
        by_mod.setdefault(p[1], []).append(p)   # p[1] = 模块
    header = ['城市', '模块', '字段', '时次', '国内值', '海外值', '差异', '是否一致', '备注', '时效分段', '写入时间']
    paths = []
    for m, pts in by_mod.items():
        path = os.path.join(out_dir, f'{name_prefix}_{m}.csv')
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(header)
            for p in pts:
                w.writerow(list(p) + [batch_time])
        paths.append(path)
    return paths


# =========================================================
# 第三部分：生成 xlsx
# =========================================================

def _weather_level_diff(cn_val, intl_val):
    """天气现象误判的量级差(并列时细分严重度:特大暴雨>暴雨>大雨),非天气返回0"""
    a = WTH_TEXTS.get(cn_val); b = WTH_TEXTS.get(intl_val)
    if not a or not b: return 0
    return abs(a['level'] - b['level'])

def aggregate_stats(allP):
    """聚合 stat：key=(field, module, period)，value 含 total/miss/clean/n/ok/
    sumdiff/maxdiff/maxcity/top/dev_counts/pair_counts。
    top: {city: (diff, str(cn), str(iv))} 仅存不一致城市的最大偏差；
    pair_counts: 天气现象 (cn, iv) 配对计数；dev_counts: 偏差值分布。"""
    stat = defaultdict(lambda: {'total': 0, 'miss': 0, 'clean': 0, 'n': 0, 'ok': 0,
                                'sumdiff': 0, 'maxdiff': 0, 'maxcity': '',
                                'top': {}, 'dev_counts': defaultdict(int),
                                'pair_counts': defaultdict(int)})
    for p in allP:
        city, module, field, ts, cnv, iv, diff, ok, note, period = p
        s = stat[(field, module, period)]
        s['total'] += 1
        if ok == '清洗剔除':
            s['clean'] += 1
            continue
        if ok in ('缺数据', ''):
            s['miss'] += 1
            continue
        s['n'] += 1
        if ok == '一致':
            s['ok'] += 1
        if isinstance(diff, (int, float)):
            s['sumdiff'] += abs(diff)
            if (abs(diff) > abs(s['maxdiff']) or
                    (abs(diff) == abs(s['maxdiff']) and
                     CITY_RANK.get(city, 9999) > CITY_RANK.get(s['maxcity'], 9999))):
                s['maxdiff'] = diff
                s['maxcity'] = city
            if ok != '一致':
                if city not in s['top'] or abs(diff) > abs(s['top'][city][0]):
                    s['top'][city] = (diff, str(cnv or ''), str(iv or ''))
            s['dev_counts'][int(diff)] += 1
        if '天气现象' in field and cnv is not None and iv is not None:
            s['pair_counts'][(str(cnv), str(iv))] += 1
    return stat

def gen_xlsx(allP, title, xlsx_path, extra_notes=None):
    """生成一致性比对报告 xlsx
    extra_notes: 可选，追加到「说明」sheet 末尾的额外说明（如均值报告的采样信息），默认 None 不追加"""
    # 从title解析拉取次数(如"阈值口径(47次均值)"或"阈值口径(47次拉取逐条)")
    import re
    _m = re.search(r'(\d+)次(均值|拉取)', title)
    _avg_n = int(_m.group(1)) if _m else 0
    _is_direct = bool(_m and _m.group(2) == '拉取')  # 逐次比对模式
    _cities = len(set(p[0] for p in allP))
    def _cnt(n): return f'{n}次' if _is_direct else (f'{n}/{_cities*_avg_n}' if _avg_n > 1 else f'{n}次')

    # ---------- Sheet1 总结 ----------
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '总结'
    # 统计逻辑不变, 但数据明细sheet不写了(有CSV留底即可)
    stat = defaultdict(lambda: {'total': 0, 'miss': 0, 'clean': 0, 'n': 0, 'ok': 0, 'sumdiff': 0, 'maxdiff': 0, 'maxcity': '', 'top': {}, 'dev_counts': defaultdict(int), 'pair_counts': defaultdict(int)})
    for p in allP:
        city, module, field, ts, cnv, iv, diff, ok, note, period = p
        s = stat[(field, module, period)]
        s['total'] += 1
        if ok == '清洗剔除':
            s['clean'] += 1; continue
        if ok in ('缺数据', ''):
            s['miss'] += 1; continue
        s['n'] += 1
        if ok == '一致': s['ok'] += 1
        if isinstance(diff, (int, float)):
            s['sumdiff'] += abs(diff)
            # 最大偏差城市: 偏差更大则替换; 偏差相同(并列)取更偏远的城市
            # (和「前五偏差城市」口径一致, 避免并列时被 CSV 顺序靠前的北京占位)
            if abs(diff) > abs(s['maxdiff']) or \
               (abs(diff) == abs(s['maxdiff']) and CITY_RANK.get(city, 9999) > CITY_RANK.get(s['maxcity'], 9999)):
                s['maxdiff'] = diff; s['maxcity'] = city
            if ok != '一致' and (city not in s['top'] or abs(diff) > abs(s['top'][city][0])):
                s['top'][city] = (diff, str(cnv or ''), str(iv or ''))
            s['dev_counts'][int(diff)] += 1
        # 天气现象字段：统计 CN→INTL 配对频次
        if '天气现象' in field and cnv is not None and iv is not None:
            s['pair_counts'][(str(cnv), str(iv))] += 1

    H2 = ['字段', '模块', '时效', '总数据', '缺数据(已排除)', '清洗剔除(已排除)', '有效样本', '一致数', '一致率', '平均偏差', '最大偏差', '最大偏差城市']
    ws.append(H2)
    for c in range(1, len(H2)+1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    PERIOD_ORDER = {p[0]: i for i, p in enumerate(PERIODS_24H)}
    for module in MODULES.keys():
        items = sorted(stat.items(), key=lambda x: (
            list(MODULES.keys()).index(x[0][1]) if x[0][1] in MODULES else 99,
            x[0][0],
            PERIOD_ORDER.get(x[0][2], 99)
        ))
        for (field, m, period), s in items:
            if m != module: continue
            rate = f"{s['ok']/s['n']*100:.1f}%" if s['n'] else '0'

            # 天气现象字段：平均偏差和最大偏差写"-"（与参考版一致）
            if '天气现象' in field:
                avg_display = '-'
                max_display = '-'
            else:
                avg_display = round(s['sumdiff'] / s['n'], 2) if s['n'] else ''
                max_display = s['maxdiff']
            # 数值字段加单位
            if '天气现象' not in field and isinstance(avg_display, (int, float)):
                u = ''
                if '温度' in field or '体感' in field: u = '℃'
                elif '湿度' in field: u = '%'
                elif '风速' in field: u = 'm/s'
                elif '气压' in field: u = 'hPa'
                elif '降水概率' in field: u = '%'
                elif '降水' in field: u = 'mm'
                avg_display = f'{avg_display}{u}'
                if isinstance(max_display, (int, float)):
                    max_display = f'{max_display}{u}'

            ws.append([field, m, period, s['total'], s['miss'], s['clean'], s['n'], s['ok'], rate, avg_display, max_display, s['maxcity']])
    for col, w in zip('ABCDEFGHIJKL', [16, 10, 14, 8, 14, 16, 10, 8, 10, 10, 10, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'

    # ---------- Sheet 天气TOP5不一致对 ----------
    ws_pairs = wb.create_sheet('天气TOP对')
    PH = ['模块', '字段', '时效', '国内天气', '国际天气', '次数']
    ws_pairs.append(PH)
    for c in range(1, len(PH)+1):
        ws_pairs.cell(row=1, column=c).font = Font(bold=True)
    for (field, m, period), s in stat.items():
        if '天气现象' not in field or not s.get('pair_counts'):
            continue
        sorted_pairs = sorted(s['pair_counts'].items(), key=lambda x: -x[1])
        # 仅统计"大类不一致"的组合(与8大类判定口径一致), 同大类文字不同(如小雨→中雨)不列为不一致对
        mismatch = []
        for (cn_val, intl_val), cnt in sorted_pairs:
            a = WTH_TEXTS.get(cn_val)
            b = WTH_TEXTS.get(intl_val)
            if a is not None and b is not None and a['cat'] != b['cat']:
                mismatch.append(((cn_val, intl_val), cnt))
        for (cn_val, intl_val), cnt in mismatch[:5]:
            ws_pairs.append([m, field, period, cn_val, intl_val, cnt])
    for col, w in zip('ABCDEF', [10, 16, 14, 14, 14, 8]):
        ws_pairs.column_dimensions[col].width = w

    # ---------- Sheet 前五偏差城市 ----------
    ws_top = wb.create_sheet('前五偏差城市')
    H3 = ['模块', '字段', '时效', '排名', '城市', 'CN→INTL', '偏差']
    ws_top.append(H3)
    for c in range(1, len(H3) + 1):
        ws_top.cell(row=1, column=c).font = Font(bold=True)
    PERIOD_ORDER_T = {p[0]: i for i, p in enumerate(PERIODS_24H)}
    for module in MODULES.keys():
        items = sorted(stat.items(), key=lambda x: (
            list(MODULES.keys()).index(x[0][1]) if x[0][1] in MODULES else 99,
            x[0][0],
            PERIOD_ORDER_T.get(x[0][2], 99)
        ))
        for (field, m, period), s in items:
            if m != module: continue
            top_items = list(s['top'].items())
            top_sorted = sorted(top_items, key=lambda x: (
                -abs(x[1][0]),
                -_weather_level_diff(x[1][1], x[1][2]),
                -CITY_RANK.get(x[0], 9999)
            ))[:5]

            for rank, (city, val) in enumerate(top_sorted, 1):
                d, cn_val, intl_val = val
                if cn_val and intl_val and '天气现象' in field:
                    pair_str = f'国内{cn_val}→国外{intl_val}'
                else:
                    pair_str = ''
                u = ''
                if '天气现象' not in field:
                    if '温度' in field or '体感' in field: u = '℃'
                    elif '湿度' in field: u = '%'
                    elif '风速' in field: u = 'm/s'
                    elif '气压' in field: u = 'hPa'
                    elif '降水概率' in field: u = '%'
                    elif '降水' in field: u = 'mm'
                d_val = round(d, 2)
                ws_top.append([m, field, period or '', rank, city, pair_str, f'{d_val}{u}' if u else d_val])
    for col, w in zip('ABCDEFG', [10, 16, 14, 6, 14, 16, 10]):
        ws_top.column_dimensions[col].width = w
    ws_top.freeze_panes = 'A2'

    # ---------- Sheet3 说明 ----------
    ws3 = wb.create_sheet('说明')
    notes = [f'一致性比对报告 — {title}', '', f'配置文件: compare_config.yaml', '']
    notes.append(f'报告生成: {len(set(p[0] for p in allP))}个城市, {len(allP)}个数据点')
    notes.append(f'比对口径: {title}')
    notes.append('')
    if '阈值' in title:
        notes.append('一、一致判定阈值(来自配置):')
        for k, v in THRESHOLDS.items():
            notes.append(f'  {k} |差|≤{v}')
        notes.append('  天气现象 按语义映射比对(详见天气映射说明)')
        notes.append('  缺数据 标"缺数据",不计入一致率分母')
        notes.append('')
    notes.append('二、风速换算(来自配置):')
    notes.append(f"  enabled={WIND_CFG.get('enabled')}, divisor={WIND_CFG.get('divisor')} (海外km/h÷3.6→m/s)")
    notes.append('')
    notes.append('三、数据清洗(物理合理范围, 来自配置 data_cleaning):')
    notes.append(f"  enabled={CLEAN_ENABLED}, 超范围的值标\"清洗剔除\"并排除出一致率分母")
    if CLEAN_ENABLED and CLEAN_RANGES:
        for k, r in CLEAN_RANGES.items():
            notes.append(f"  {k}: {r.get('min')}~{r.get('max')}{r.get('unit','')} ({r.get('basis','')})")
    else:
        notes.append('  (未配置清洗范围, 不剔除任何值)')
    notes.append('')
    notes.append('四、24小时时效分段: ' + ' / '.join(p[0] for p in PERIODS_24H))
    notes.append('')
    notes.append('五、时次对齐: 24h按predict_time, 15天按predict_date')
    notes.append('')
    notes.append('六、天气现象语义映射比对规则（五分制评分）：')
    notes.append('  将国内外中文天气文字统一映射到(大类, 量级, 是否高影响)')
    notes.append('  评分→偏差规则:')
    notes.append('    5分 主天气一致+量级一致                     → 完全匹配')
    notes.append('    4分 主天气一致+量级差1级                     → 轻微量级偏差')
    notes.append('    3分 主天气不一致(晴↔多云等)，均非高影响       → 主天气不一致')
    notes.append('    2分 主天气错判(量级差≥2)，或降水vs非降水     → 明显偏差')
    notes.append('    1分 涉及高影响天气+同大类                   → 高影响偏差')
    notes.append('    0分 高影响天气+不同大类                     → 高影响漏报/错判')
    notes.append('  一致判定: 仅5分(偏差=0)算"一致"，其余全算"不一致"')
    notes.append('  高影响天气: 大雨/暴雨/大暴雨/特大暴雨/大雪/暴雪/雷暴/冰雹')
    notes.append('  完整映射对照表请见 compare_config.yaml → weather_mapping')
    notes.append('')
    notes.append('注: 修改 compare_config.yaml 后重跑即可更新阈值口径')
    for n in notes:
        ws3.append([n])
    if extra_notes:
        ws3.append([''])
        for n in extra_notes:
            ws3.append([n])
    ws3['A1'].font = Font(bold=True, size=13)

    wb.save(xlsx_path)
    return len(allP)

# =========================================================
# 主流程
# =========================================================

def load_cities():
    """读城市列表（去重），并设置 CITY_RANK（CSV 顺序，偏差相同时偏远地区优先）。
    返回 [(name, lon, lat), ...]。供一次性脚本和定时脚本复用"""
    global CITY_RANK
    cities = []
    seen = set()
    with open(CITY_CSV, 'r', encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            lon = r['Flon'].strip(); lat = r['Flat'].strip()
            key = (lon, lat)
            if key in seen: continue
            seen.add(key)
            cities.append((r['Fcityname_cn'].strip(), lon, lat))
    # 城市序号：CSV 顺序（北京、天津、上海...漠河、阿勒泰），偏差相同时偏远地区优先
    CITY_RANK = {name: i for i, (name, *_rest) in enumerate(cities)}
    return cities

def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    # 1. 读城市列表（去重）+ 设置偏远优先序号
    cities = load_cities()
    print(f"共 {len(cities)} 个城市, 开始请求实时 API...")

    # 2. 逐个城市拉数据 + 比对
    allP_threshold = []
    ok_count = 0

    for idx, (name, lon, lat) in enumerate(cities, 1):
        cn_data, in_data = fetch_city(name, lon, lat)

        if cn_data is None:
            print(f"  [{idx}/{len(cities)}] ⏭️ {name} 国内接口失败, 跳过")
            continue
        if in_data is None:
            print(f"  [{idx}/{len(cities)}] ⏭️ {name} 国际接口失败, 跳过")
            continue

        allP_threshold += build_points(name, cn_data, in_data)
        ok_count += 1

        if idx % 10 == 0 or idx == len(cities):
            print(f"  [{idx}/{len(cities)}] {name} ✅")

    print(f"\n请求完成: 成功 {ok_count} 个城市, 跳过 {len(cities) - ok_count} 个")

    if not allP_threshold:
        print("❌ 无有效数据, 不生成报告")
        return

    # 3. 生成阈值口径报告
    n2 = gen_xlsx(allP_threshold, '阈值口径', XLSX_THRESHOLD)
    print(f"✅ 阈值口径报告: {XLSX_THRESHOLD}")
    print(f"   数据点: {n2}")

    # 4. 导出数据明细CSV(按模块分文件, 每行带写入时间标识批次; 替代原始JSON作为证据)
    csv_paths = write_detail_csv(allP_threshold, OUT_DIR, f'数据明细_{TIMESTAMP}')
    print(f"✅ 数据明细CSV(按模块分文件, 共{len(csv_paths)}个):")
    for p in csv_paths:
        print(f"   {os.path.basename(p)}")

    # 5. 快速打印阈值口径一致率
    print(f"\n{'='*60}")
    print(f"阈值口径 — 各字段一致率速览")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(XLSX_THRESHOLD)
    ws = wb['总结']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and not row[2]:  # 无时效分段的汇总行（实况/15天/AQI）
            print(f"  {str(row[0]):14s} {str(row[1]):8s}  一致率: {str(row[8]):>7s}  平均偏差: {str(row[9]):>8s}")

if __name__ == '__main__':
    main()
