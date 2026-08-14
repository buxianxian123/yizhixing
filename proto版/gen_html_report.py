#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成气派 HTML 一致性比对报告（单文件离线 / ECharts / 全局筛选+每表排序 / 浅色商务风）

读 xlsx(总结) + CSV明细 -> 内联 echarts.min.js -> 单文件 HTML
- 全局筛选器统一联动所有图表（模块/字段/时效/地区三级联动）
- 每个表自带排序控件，无"跟随全局/独立"切换
- 大表独占一排、小表两两一排，字号放大
- 分模块 TOP5 偏差城市表（6张，md 格式：行字段列TOP1-5城市）

运行:
  python3 gen_html_report.py              # 默认取最新阈值口径 xlsx + 对应CSV
  python3 gen_html_report.py <xlsx路径>   # 指定 xlsx
"""
import os, sys, json, glob, datetime, csv
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.join(SCRIPT_DIR, '..', 'data')
OUT_DIR = os.path.join(BASE, '比对结果')
ECHARTS_PATH = os.path.join(SCRIPT_DIR, 'echarts.min.js')
CONFIG_PATH = os.environ.get('CONFIG_PATH') or os.path.join(SCRIPT_DIR, 'compare_config.yaml')
TIMESTAMP = datetime.datetime.now().strftime('%Y%m%d_%H%M')

def parse_window_end(xlsx):
    """从xlsx文件名解析采样窗口结束时间(数据截止), 如 2026-07-29 10:36; 无窗口返回当前时间"""
    import re
    m = re.search(r'(\d{8}_\d{4,6})-(\d{8}_\d{4,6})', os.path.basename(xlsx))
    if m:
        s = m.group(2)
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]} {s[9:11]}:{s[11:13]}"
    return datetime.datetime.now().strftime('%Y-%m-%d %H:%M')


MODULES_DISPLAY = {'实况': '实况', '24小时': '24小时', '15天': '15天', 'AQI模块': 'AQI'}
PERIODS = ['短时效(1-6h)', '中时效(7-12h)', '长时效(13-24h)']

# 城市 -> [国, 省/国] 映射（70 城，含港澳台归中国、5 个海外）
CITY_REGION = {
    '北京市': ['中国', '北京'], '上海市': ['中国', '上海'], '天津市': ['中国', '天津'], '重庆市': ['中国', '重庆'],
    '唐山市': ['中国', '河北'], '石家庄市': ['中国', '河北'],
    '太原市': ['中国', '山西'],
    '大连市': ['中国', '辽宁'], '沈阳市': ['中国', '辽宁'],
    '长春市': ['中国', '吉林'], '延吉市': ['中国', '吉林'],
    '哈尔滨市': ['中国', '黑龙江'], '齐齐哈尔市': ['中国', '黑龙江'], '佳木斯市': ['中国', '黑龙江'], '漠河市': ['中国', '黑龙江'],
    '南京市': ['中国', '江苏'], '苏州市': ['中国', '江苏'], '无锡市': ['中国', '江苏'], '徐州市': ['中国', '江苏'],
    '杭州市': ['中国', '浙江'], '宁波市': ['中国', '浙江'], '温州市': ['中国', '浙江'],
    '合肥市': ['中国', '安徽'],
    '福州市': ['中国', '福建'], '厦门市': ['中国', '福建'],
    '南昌市': ['中国', '江西'],
    '青岛市': ['中国', '山东'], '济南市': ['中国', '山东'], '烟台市': ['中国', '山东'],
    '郑州市': ['中国', '河南'], '洛阳市': ['中国', '河南'],
    '武汉市': ['中国', '湖北'],
    '长沙市': ['中国', '湖南'],
    '广州市': ['中国', '广东'], '深圳市': ['中国', '广东'], '珠海市': ['中国', '广东'], '东莞市': ['中国', '广东'], '佛山市': ['中国', '广东'],
    '南宁市': ['中国', '广西'],
    '三亚市': ['中国', '海南'], '海口市': ['中国', '海南'],
    '成都市': ['中国', '四川'],
    '贵阳市': ['中国', '贵州'],
    '昆明市': ['中国', '云南'], '丽江市': ['中国', '云南'],
    '拉萨市': ['中国', '西藏'],
    '西安市': ['中国', '陕西'],
    '兰州市': ['中国', '甘肃'], '敦煌市': ['中国', '甘肃'],
    '西宁市': ['中国', '青海'],
    '呼和浩特市': ['中国', '内蒙古'],
    '乌鲁木齐市': ['中国', '新疆'], '喀什市': ['中国', '新疆'], '喀什地区': ['中国', '新疆'],
    '库尔勒市': ['中国', '新疆'], '吐鲁番地区': ['中国', '新疆'], '阿勒泰市': ['中国', '新疆'], '阿勒泰地区': ['中国', '新疆'],
    '香港特别行政区': ['中国', '香港'], '澳门特别行政区': ['中国', '澳门'], '台北市': ['中国', '台湾'],
    '吉隆坡': ['海外', '马来西亚'], '新加坡': ['海外', '新加坡'], '多伦多': ['海外', '加拿大'], '柏林': ['海外', '德国'], '洛杉矶县': ['海外', '美国'],
}


# =========================================================
# 数据读取
# =========================================================

def find_inputs(xlsx_arg=None):
    if xlsx_arg:
        xlsx = xlsx_arg
    else:
        fs = glob.glob(os.path.join(OUT_DIR, '**', '一致性比对报告_均值_阈值口径_*.xlsx'), recursive=True)
        if not fs:
            raise SystemExit('❌ 未找到阈值口径 xlsx，请先跑 reformat_threshold.py / scheduled_compare.py')
        xlsx = max(fs, key=os.path.getmtime)
    # csv 与 xlsx 同源（同tag），递归匹配子目录（定时任务报告在 阈值口径报告/<日期>/、数据明细CSV/<日期>/）
    tag = os.path.basename(xlsx).replace('一致性比对报告_均值_阈值口径_', '').replace('.xlsx', '')
    cs = glob.glob(os.path.join(OUT_DIR, '**', f'数据明细_均值_{tag}.csv'), recursive=True)
    csv_path = cs[0] if cs else None
    return xlsx, csv_path


def read_echarts():
    if os.path.exists(ECHARTS_PATH):
        with open(ECHARTS_PATH, encoding='utf-8') as f:
            return f.read()
    print('本地未找到 echarts.min.js, 从 CDN 下载...')
    import urllib.request
    urllib.request.urlretrieve(
        'https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js', ECHARTS_PATH)
    with open(ECHARTS_PATH, encoding='utf-8') as f:
        return f.read()


def read_summary(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb['总结']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    out = []
    for r in rows[1:]:
        if r is None or r[0] is None:
            continue
        r = list(r) + [None] * (len(headers) - len(r))
        d = dict(zip(headers, r))
        out.append({
            'field': d['字段'], 'module': d['模块'], 'period': d['时效'] or '',
            'total': d['总数据'], 'miss': d['缺数据(已排除)'], 'valid': d['有效样本'],
            'ok': d['一致数'], 'rate': d['一致率'], 'avgDev': d['平均偏差'],
            'maxDev': d['最大偏差'], 'maxCity': d['最大偏差城市']
        })
    return out


def read_detail(csv_path):
    out = []
    with open(csv_path, encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            out.append([
                r['城市'], r['模块'], r['字段'], r['时次'],
                r['国内值'], r['海外值'], r['差异'], r['是否一致'], r['时效分段'] or ''
            ])
    return out


def read_thresholds():
    try:
        import yaml
        cfg = yaml.safe_load(open(CONFIG_PATH, encoding='utf-8'))
        return cfg.get('thresholds', {})
    except Exception:
        return {}


def read_weather_texts():
    """读天气现象语义映射(文字->{level,cat,hi})，供前端TOP5按量级差细分排序"""
    try:
        import yaml
        cfg = yaml.safe_load(open(CONFIG_PATH, encoding='utf-8'))
        return cfg.get('weather_mapping', {}).get('texts', {})
    except Exception:
        return {}


def load_city_rank():
    """读城市CSV返回 {城市名:序号}(CSV顺序,偏远靠后),用于并列偏差时偏远优先,与reformat_threshold口径一致"""
    city_csv = os.path.join(BASE, '天气一致性测试城市_热门城市筛选.csv')
    rank = {}
    seen = set()
    try:
        with open(city_csv, encoding='utf-8-sig') as f:
            for r in csv.DictReader(f):
                name = r['Fcityname_cn'].strip()
                if name in seen:
                    continue
                seen.add(name)
                rank[name] = len(rank)
    except Exception:
        pass
    return rank


# =========================================================
# HTML 生成
# =========================================================

def build_html(data_json, echarts_js, meta, thresholds, rain_th):
    echarts_js = echarts_js.replace('</script>', '<\\/script>')
    data_json = data_json.replace('</script>', '<\\/script>')

    th_rows = ''.join(
        f'<tr><td>{k}</td><td>≤ {v}</td></tr>' for k, v in thresholds.items()
    ) or '<tr><td colspan="2">（读配置失败）</td></tr>'

    rain_rows = ''.join(
        f'<tr><td>{n}</td><td>{"∞" if th in ("~", None) else th}</td></tr>'
        for n, th in rain_th.items()
    ) if rain_th else '<tr><td colspan="2">（无降水量配置）</td></tr>'

    html = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>墨迹天气 国内/国际 数据一致性比对报告</title>
<style>
:root{
  --brand:#2563eb; --brand-d:#1d4ed8; --brand-l:#eff6ff;
  --ink:#0f172a; --ink2:#475569; --ink3:#94a3b8;
  --line:#e5e7eb; --bg:#f5f7fa; --card:#ffffff;
  --ok:#16a34a; --bad:#dc2626; --warn:#f59e0b;
  --shadow:0 1px 3px rgba(15,23,42,.06),0 1px 2px rgba(15,23,42,.04);
  --shadow-h:0 4px 16px rgba(15,23,42,.10);
  /* TOP排名特殊配色：5档从浅黄渐变到白，第一档也不深 */
  --top1:#fef9c3; --top2:#fefce8; --top3:#fffdf2; --top4:#fffef8; --top5:#fffefa;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif;
  background:var(--bg);color:var(--ink);line-height:1.5;-webkit-font-smoothing:antialiased}
.wrap{max-width:1520px;margin:0 auto;padding:24px 28px 60px}
.header{background:linear-gradient(135deg,#1e3a8a 0%,#2563eb 100%);color:#fff;border-radius:14px;
  padding:26px 30px;box-shadow:var(--shadow-h);margin-bottom:18px}
.header h1{font-size:26px;font-weight:700;letter-spacing:.5px}
.header .sub{margin-top:10px;font-size:14px;opacity:.9;display:flex;gap:22px;flex-wrap:wrap}
.header .sub b{font-weight:600}
/* 筛选器：控件居中平铺，说明放下方居中 */
.filters{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:14px 18px;
  margin-bottom:18px;box-shadow:var(--shadow);display:flex;flex-direction:column;gap:10px}
.filters-main{display:flex;align-items:center;gap:18px;flex-wrap:wrap;justify-content:center}
.filters .fg{display:flex;align-items:center;gap:8px}
.filters .ftitle{font-size:13px;font-weight:600;color:var(--ink2)}
.filters .sel{display:inline-flex;align-items:center;gap:5px}
.filters .sel>span{font-size:12px;color:var(--ink3);font-weight:500}
.filters select,.filters input{height:34px;border:1px solid var(--line);border-radius:8px;
  padding:0 10px;font-size:13px;background:#fff;color:var(--ink);outline:none;transition:.15s}
.filters select{min-width:96px}
.filters select:focus,.filters input:focus{border-color:var(--brand);box-shadow:0 0 0 3px var(--brand-l)}
.filters .reset{height:34px;padding:0 14px;border:1px solid var(--line);
  background:#fff;border-radius:8px;font-size:13px;cursor:pointer;color:var(--ink2);font-weight:500;transition:.15s}
.filters .reset:hover{border-color:var(--brand);color:var(--brand)}
.filters .arrow{color:var(--ink3);font-size:13px}
.filters .filter-help{display:flex;justify-content:center;gap:34px;flex-wrap:wrap;
  font-size:12px;color:var(--ink3);line-height:1.6;border-top:1px dashed var(--line);padding-top:10px}
.filters .filter-help b{color:var(--ink2);font-weight:600;margin-right:3px}
/* KPI */
.kpis{display:grid;grid-template-columns:1fr 1fr .7fr .7fr 1fr 1.7fr;gap:14px;margin-bottom:18px}
.kpi{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:18px 20px;
  box-shadow:var(--shadow);position:relative;overflow:hidden}
.kpi::before{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--brand)}
.kpi .label{font-size:13px;color:var(--ink3);font-weight:500}
.kpi .value{font-size:32px;font-weight:700;color:var(--ink);margin-top:6px;line-height:1.1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.kpi.kpi-name .value{font-size:22px}
.kpi .value small{font-size:16px;font-weight:600;color:var(--ink2);margin-left:2px}
.kpi .extra{font-size:13px;color:var(--ink2);margin-top:4px}
.kpi.warn::before{background:var(--bad)}
.kpi.warn .value{color:var(--bad)}
/* 图表网格 */
.grid{display:grid;grid-template-columns:repeat(2,1fr);gap:16px;margin-bottom:16px}
.grid .span2{grid-column:span 2}
.card{background:var(--card);border:1px solid var(--line);border-radius:12px;box-shadow:var(--shadow);
  overflow:hidden;display:flex;flex-direction:column;margin-bottom:0}
.card-head{display:flex;align-items:center;justify-content:space-between;padding:14px 18px;
  border-bottom:1px solid var(--line);background:#fafbfc}
.card-title{font-size:16px;font-weight:600;color:var(--ink)}
.card-ctrl{display:flex;align-items:center;gap:8px}
.ctl-btn{height:30px;padding:0 12px;border:1px solid var(--line);background:#fff;border-radius:6px;
  font-size:13px;cursor:pointer;color:var(--ink2);transition:.15s;font-weight:500}
.ctl-btn:hover{border-color:var(--brand);color:var(--brand)}
.ctl-btn.active{background:var(--brand-l);border-color:var(--brand);color:var(--brand-d)}
.ctl-select{height:30px;padding:0 8px;border:1px solid var(--line);background:#fff;border-radius:6px;
  font-size:13px;cursor:pointer;color:var(--ink2);outline:none;transition:.15s}
.ctl-select:hover{border-color:var(--brand)}
.ctl-select:focus{border-color:var(--brand);box-shadow:0 0 0 3px var(--brand-l)}
.chart-body{width:100%;padding:14px}
.chart-body.xl{height:520px}
.chart-body.tall{height:440px}
.chart-body.mid{height:360px}
.chart-body.short{height:300px}
/* 24h趋势小多图网格 */
.trend24-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;padding:14px}
.trend24-cell{background:#fafbfc;border:1px solid var(--line);border-radius:8px;padding:8px 8px 4px;transition:.15s}
.trend24-cell:hover{box-shadow:0 2px 8px rgba(15,23,42,.08)}
.trend24-title{font-size:12px;font-weight:600;color:var(--ink2);padding:0 2px 4px;text-align:center}
.trend24-chart{height:200px;width:100%}
/* 表格全面优化 */
.tbl-wrap{padding:16px 20px 20px;overflow:auto}
.tbl-tools{display:flex;align-items:center;gap:12px;margin-bottom:12px;font-size:13px;color:var(--ink2)}
.tbl-tools input{height:32px;border:1px solid var(--line);border-radius:6px;padding:0 10px;font-size:13px;width:240px}
.tbl-tools input:focus{outline:none;border-color:var(--brand)}
.tbl-tools .pager{margin-left:auto;display:flex;align-items:center;gap:6px}
.tbl-tools .pager button{height:32px;padding:0 12px;border:1px solid var(--line);background:#fff;
  border-radius:6px;cursor:pointer;font-size:13px;transition:.15s}
.tbl-tools .pager button:hover{border-color:var(--brand);color:var(--brand)}
.tbl-tools .pager button:disabled{opacity:.4;cursor:not-allowed}
table{width:100%;border-collapse:collapse;font-size:14px}
thead th{background:#f1f5f9;color:var(--ink2);font-weight:600;padding:12px 14px;text-align:left;
  border-bottom:2px solid var(--line);white-space:nowrap;position:sticky;top:0}
thead th.sortable{cursor:pointer}
thead th.sortable:hover{color:var(--brand)}
tbody td{padding:12px 14px;border-bottom:1px solid #f1f5f9;vertical-align:middle}
tbody tr:nth-child(even){background:#fafbfc}
tbody tr:hover{background:var(--brand-l)}
/* 列宽和对齐优化 */
.fname{font-weight:600;color:var(--ink);min-width:120px}
td.city{min-width:100px}
td.num{text-align:right;font-variant-numeric:tabular-nums;min-width:80px}
td small{color:var(--ink3);font-size:12px;font-weight:500}
td .dev{font-weight:600}
.dev-h{color:var(--bad)} .dev-m{color:var(--warn)} .dev-l{color:var(--ink)}
/* TOP排名特殊样式：5档浅黄渐变到白 */
td.top1{background:var(--top1);font-weight:600}
td.top2{background:var(--top2)}
td.top3{background:var(--top3)}
td.top4{background:var(--top4)}
td.top5{background:var(--top5)}
.tag{display:inline-block;padding:2px 10px;border-radius:10px;font-size:12px;font-weight:600}
.tag.ok{background:#dcfce7;color:#15803d}
.tag.bad{background:#fee2e2;color:#b91c1c}
.tag.miss{background:#f1f5f9;color:#64748b}
/* 明细表按需查询 */
.detail-query{padding:16px 20px;border-bottom:1px solid var(--line)}
.detail-query .dq-row{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.detail-query .ftitle{font-size:13px;font-weight:600;color:var(--ink2)}
.detail-query select,.detail-query input{height:32px;border:1px solid var(--line);border-radius:6px;
  padding:0 10px;font-size:13px;background:#fff;color:var(--ink);outline:none;transition:.15s}
.detail-query select{min-width:108px}
.detail-query input{width:200px}
.detail-query select:focus,.detail-query input:focus{border-color:var(--brand);box-shadow:0 0 0 3px var(--brand-l)}
.detail-query .dq-tip{margin-top:10px;font-size:12px;color:var(--ink3);line-height:1.6}
.detail-query .dq-tip.err{color:var(--bad);font-weight:500}
.detail-result .tbl-wrap{padding:16px 20px 20px}
.hidden{display:none!important}
/* 左图+右栏布局 */
.split-layout{display:flex;align-items:stretch}
.split-layout .chart-body{flex:1;min-width:0}
.side-list{width:190px;flex-shrink:0;border-left:1px solid var(--line);padding:14px 12px;overflow:auto;background:#fafbfc}
.side-list .sl-title{font-size:12px;color:var(--ink3);font-weight:600;margin:4px 0 6px;letter-spacing:.5px}
.side-list .sl-item{display:flex;justify-content:space-between;align-items:center;padding:7px 10px;border-radius:6px;
  cursor:pointer;font-size:13px;color:var(--ink2);transition:.12s;margin-bottom:2px}
.side-list .sl-item:hover{background:#fff}
.side-list .sl-item.active{background:var(--brand-l);color:var(--brand-d);font-weight:600}
.side-list .sl-item b{font-variant-numeric:tabular-nums}
.side-list .sl-item.bad b{color:var(--bad)}
.side-list .sl-item.ok b{color:var(--ok)}
.side-list .sl-empty{font-size:12px;color:var(--ink3);padding:6px 10px}
/* 卡片副标题 + 解读条 */
.card-sub{font-size:12px;color:var(--ink3);font-weight:400;margin-left:8px}
.card-note{padding:11px 18px;border-top:1px solid var(--line);background:#fafbfc;font-size:13px;color:var(--ink2);line-height:1.7}
.card-note b{color:var(--ink);font-weight:600}
.sankey-tbl-wrap{padding:4px 18px 16px;border-top:1px solid var(--line)}
.sankey-twin{display:flex;gap:18px}
.sankey-col{flex:1;min-width:0}
.sankey-col .sl-title{font-size:12px;color:var(--ink3);font-weight:600;margin:4px 0 6px}
.sankey-tbl-wrap table{font-size:13px;margin:0}
.sankey-tbl-wrap thead th{position:static;background:#f8fafc}
.sankey-tbl-wrap tbody td{padding:8px 12px}
.sankey-tbl-wrap .rank{width:36px;color:var(--ink3);font-variant-numeric:tabular-nums}
/* 底部说明 */
.notes{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:20px 24px;
  box-shadow:var(--shadow);margin-top:4px}
.notes h3{font-size:17px;font-weight:600;margin-bottom:12px;color:var(--ink)}
.notes h4{font-size:14px;font-weight:600;margin:16px 0 8px;color:var(--ink2)}
.notes p,.notes li{font-size:14px;color:var(--ink2);line-height:1.8}
.notes ul{padding-left:22px}
.notes table{margin:8px 0;font-size:14px}
.notes thead th{cursor:default}
.notes .risk{background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:12px 16px;margin-top:8px}
.notes .risk li{color:#92400e}
@media(max-width:1100px){.kpis{grid-template-columns:repeat(3,1fr)}.grid{grid-template-columns:1fr}.grid .span2{grid-column:span 1}}
</style>
</head>
<body>
<div class="wrap">
  <div class="header">
    <h1>墨迹天气 · 国内 / 国际 数据一致性比对报告</h1>
    <div class="sub">
      <span>报告时间：<b>__TIME__</b></span>
      <span>数据源：<b>__SOURCE__</b></span>
      <span>口径：<b>__KOUJING__</b></span>
      <span>城市：<b>__CITIES__</b></span>
      <span>数据点：<b>__POINTS__</b></span>
    </div>
  </div>

  <!-- 全局筛选器：控件居中平铺 + 下方说明 -->
  <div class="filters">
    <div class="filters-main">
      <div class="fg">
        <span class="ftitle">模块维度</span>
        <label class="sel"><span>模块</span><select id="g-module" onchange="applyGlobal()"></select></label>
        <label class="sel"><span>字段</span><select id="g-field" onchange="applyGlobal()"></select></label>
        <label class="sel"><span>时效</span><select id="g-period" onchange="applyGlobal()"></select></label>
      </div>
      <div class="fg">
        <span class="ftitle">地区</span>
        <label class="sel"><span>国</span><select id="r1" onchange="onRegion1(this.value)"></select></label>
        <span class="arrow">›</span>
        <label class="sel"><span>省</span><select id="r2" onchange="onRegion2(this.value)"></select></label>
        <span class="arrow">›</span>
        <label class="sel"><span>市</span><select id="r3" onchange="onRegion3(this.value)"></select></label>
      </div>
      <button class="reset" onclick="resetGlobal()">重置全部</button>
    </div>
    <div class="filter-help">
      <span><b>模块</b>实况/24h/15天/AQI</span>
      <span><b>字段</b>温度/湿度/风速/气压/天气现象/AQI</span>
      <span><b>时效</b>短1-6h/中7-12h/长13-24h</span>
      <span><b>地区</b>国›省›市，留空即该层全部</span>
    </div>
  </div>

  <!-- KPI -->
  <div class="kpis" id="kpis"></div>

  <!-- 图表区 -->
  <div class="grid">
    <div class="card" id="card-moduleBar"><div class="card-head"><span class="card-title">模块一致率</span></div>
      <div class="chart-body mid" id="ch-moduleBar"></div></div>
    <div class="card" id="card-pie"><div class="card-head"><span class="card-title">一致分布</span></div>
      <div class="chart-body short" id="ch-pie"></div></div>

    <div class="card span2" id="card-fieldBar"><div class="card-head"><span class="card-title">字段一致率<span class="card-sub">按一致率排序</span></span>
      <div class="card-ctrl"><button class="ctl-btn" id="fb-sort" onclick="toggleFieldSort()">升降序切换</button></div></div>
      <div class="split-layout">
        <div class="chart-body xl" id="ch-fieldBar"></div>
        <div class="side-list" id="fb-module-list"></div>
      </div>
      <div class="card-note" id="fieldBar-note"></div></div>

    <div class="card span2" id="card-trend24"><div class="card-head"><span class="card-title">24小时时效准确率趋势<span class="card-sub">每字段一小图 · 短/中/长时效一致率</span></span></div>
      <div class="trend24-grid" id="trend24-grid"></div></div>
    <div class="card span2" id="card-devHist"><div class="card-head"><span class="card-title">偏差分布<span class="card-sub">点选右侧字段，各字段量级独立</span></span></div>
      <div class="split-layout">
        <div class="chart-body mid" id="ch-devHist"></div>
        <div class="side-list" id="dev-field-list"></div>
      </div></div>

    <div class="card span2" id="card-cityRank"><div class="card-head"><span class="card-title">城市一致率排行<span class="card-sub">按一致率排序</span></span>
      <div class="card-ctrl" id="city-view-ctrl">
        <button class="ctl-btn" onclick="setCityView(5,this)">5</button>
        <button class="ctl-btn" onclick="setCityView(10,this)">10</button>
        <button class="ctl-btn" onclick="setCityView(15,this)">15</button>
        <button class="ctl-btn" onclick="setCityView(20,this)">20</button>
        <button class="ctl-btn active" onclick="setCityView(25,this)">25</button>
        <button class="ctl-btn" onclick="setCityView(0,this)">全部</button>
      </div></div>
      <div class="split-layout">
        <div class="chart-body tall" id="ch-cityRank"></div>
        <div class="side-list" id="city-rank-list"></div>
      </div></div>

    <div class="card span2" id="card-sankey"><div class="card-head"><span class="card-title">天气现象误判流向<span class="card-sub">国内 › 海外，仅不一致项</span></span></div>
      <div class="chart-body tall" id="ch-sankey"></div>
      <div class="card-note" id="sankey-note"></div>
      <div class="sankey-tbl-wrap" id="sankey-table"></div></div>

    <!-- 分模块 TOP5 表（动态生成） -->
    <div class="card span2" style="background:transparent;border:none;box-shadow:none;padding:0">
      <div id="top5-zone"></div>
    </div>

    <!-- 数据明细表（按需查询加载，避免一次性渲染全量） -->
    <div class="card span2" id="card-detail">
      <div class="card-head"><span class="card-title">数据明细</span>
        <span style="font-size:12px;color:var(--ink3)">选择条件后点「查询」加载，单次上限 5000 条</span></div>
      <div class="detail-query">
        <div class="dq-row">
          <span class="ftitle">模块</span>
          <select id="d-module"></select>
          <span class="ftitle">字段</span>
          <select id="d-field"></select>
          <span class="ftitle">时效</span>
          <select id="d-period"></select>
          <input id="d-search" placeholder="🔍 搜索城市/字段/模块...">
          <button class="ctl-btn active" onclick="queryDetail()">查询</button>
          <button class="ctl-btn" id="d-collapse" onclick="collapseDetail()" style="display:none">收起</button>
        </div>
        <div class="dq-tip" id="dq-tip">提示：至少选择一个筛选维度或输入搜索词，避免全量加载。全量约 <b id="dq-total">-</b> 条</div>
      </div>
      <div class="detail-result" id="detail-result" style="display:none">
        <div class="tbl-wrap" id="ch-detail"></div>
      </div>
    </div>
  </div>

  <!-- 底部说明 -->
  <div class="notes">
    <h3>测试说明 · 阈值配置 · 风险遗留</h3>
    <h4>一、一致判定阈值（数值字段）</h4>
    <table><thead><tr><th>字段</th><th>一致判定阈值</th></tr></thead><tbody>__TH_ROWS__</tbody></table>
    <p>差异 = 国内值 − 海外值；正数表示国内 &gt; 海外，负数表示国内 &lt; 海外。缺数据不计入一致率分母。</p>
    __RAIN_SECTION__
    <h4>三、天气现象语义映射（五分制）</h4>
    <ul>
      <li><b>5分</b> 主天气一致+量级一致 -> 完全匹配（唯一算"一致"）</li>
      <li><b>4分</b> 主天气一致+量级差1级 -> 轻微量级偏差</li>
      <li><b>3分</b> 主天气不一致（晴↔多云等），均非高影响</li>
      <li><b>2分</b> 主天气错判（量级差≥2），或降水vs非降水</li>
      <li><b>1分</b> 涉及高影响天气+同大类</li>
      <li><b>0分</b> 高影响天气+不同大类 -> 高影响漏报/错判</li>
    </ul>
    <p>高影响天气：大雨/暴雨/大暴雨/特大暴雨/大雪/暴雪/雷暴/冰雹。阴/雾/霾归入多云大类。</p>
    <h4>四、风速换算</h4>
    <p>国内风速 m/s，海外 km/h，海外值 ÷3.6 换算为 m/s 后比对。</p>
    <h4>五、风险与遗留项</h4>
    <ul class="risk">
      <li>海外城市未覆盖：国内接口不支持海外城市，本次以国内城市为主，部分城市如吉隆坡等为国内接口可覆盖的海外城市</li>
      <li>AQI 一致率极低（约 4.6%），两套数据源存在系统性差异，需排查归一化逻辑</li>
      <li>气压高原系统性偏差：拉萨等高原城市偏差可达 400hPa，疑海拔/气压基准不同</li>
      <li>体感温度偏差大：15天白天一致率仅约 10%，体感计算模型差异可能是主因</li>
      <li>数据为多次拉取均值，仍有时间窗口内波动影响</li>
    </ul>
  </div>
</div>

<script>__ECHARTS__</script>
<script>
const DATA = __DATA__;
const detail = DATA.detail;
const REGION = DATA.region;        // {城市:[国,省]}
const C=0,MO=1,F=2,TS=3,CN=4,IV=5,DF=6,OK=7,PR=8;

const G = {module:'全部', field:'全部', period:'全部', cities:[]};

const MODULE_LIST = ['实况','24小时','15天','AQI模块'];
const FIELD_LIST = [...new Set(detail.map(r=>r[F]))].sort();
const PERIOD_LIST = ['短时效(1-6h)','中时效(7-12h)','长时效(13-24h)'];
// 字段展示顺序
const FIELD_ORDER = ['温度','体感温度','湿度','风速','气压','天气现象','降水量',
  '温度(最高)','温度(最低)','体感温度(白天)','体感温度(夜间)',
  '风速(白天)','风速(夜间)','天气现象(白天)','天气现象(夜间)','AQI'];

const BASE_TEXT = {color:'#334155', fontFamily:'inherit', fontSize:13};
const GRID = {left:56, right:28, top:40, bottom:40};

// 地区树
const REGION_TREE = (function(){
  const t={};
  for(const city in REGION){
    const [c,p]=REGION[city];
    if(!t[c])t[c]={};
    if(!t[c][p])t[c][p]=[];
    t[c][p].push(city);
  }
  return t;
})();

function num(v){ if(v===null||v===undefined||v==='')return null; const n=parseFloat(v); return isNaN(n)?null:n; }
function isWeather(f){ return f && f.indexOf('天气现象')>=0; }
function modShow(m){ return m==='AQI模块'?'AQI':m; }
function devClass(d){ const a=Math.abs(d); return a>10?'dev-h':a>3?'dev-m':'dev-l'; }

function filterRows(f){
  const mc=f.module, fc=f.field, pc=f.period, cs=f.cities||[];
  return detail.filter(r =>
    (mc==='全部' || r[MO]===mc) &&
    (fc==='全部' || r[F]===fc) &&
    (pc==='全部' || r[PR]===pc) &&
    (cs.length===0 || cs.indexOf(r[C])>=0)
  );
}
function aggRate(rows){
  let ok=0,bad=0,miss=0;
  for(const r of rows){
    if(r[OK]==='一致')ok++; else if(r[OK]==='不一致')bad++; else miss++;
  }
  const valid=ok+bad;
  return {ok,bad,miss,valid,rate:valid?ok/valid*100:0};
}

// ====== 图实例 ======
const EC = {};
function initCharts(){
  for(const id of ['moduleBar','fieldBar','pie','cityRank','devHist','sankey']){
    EC[id] = echarts.init(document.getElementById('ch-'+id));
  }
  window.addEventListener('resize', ()=>{
    for(const id in EC) EC[id].resize();
    TREND24_CELLS.forEach(c=>c.resize());
  });
}

// ====== KPI ======
function drawKPI(){
  const rows = filterRows(G);
  const a = aggRate(rows);
  const cities = new Set(rows.map(r=>r[C])).size;
  const mods = new Set(rows.map(r=>r[MO])).size;
  const grp={};
  for(const r of rows){ const k=r[F]+'@'+r[MO]; if(!grp[k])grp[k]={f:r[F],m:r[MO],rows:[]}; grp[k].rows.push(r); }
  let weakest=null;
  for(const k in grp){ const g=grp[k]; const ar=aggRate(g.rows);
    if(ar.valid>=10 && (!weakest || ar.rate<weakest.rate)) weakest={f:g.f,m:g.m,rate:ar.rate}; }
  // 实况最大偏差城市：按 |偏差|/阈值 比例最大，并列优先偏远（cityRank大的）
  const TH=DATA.thresholds||{}, CR=DATA.cityRank||{};
  const getTh=f=>{for(const k in TH){if(f.indexOf(k)>=0||k.indexOf(f)>=0)return TH[k];}return null;};
  let bestRatio=-1, bestCity='-', bestField='', bestDev=0;
  for(const r of rows){
    if(r[MO]!=='实况'||isWeather(r[F])) continue;
    const d=num(r[DF]); if(d===null) continue;
    const th=getTh(r[F]); if(!th) continue;
    const ratio=Math.abs(d)/th;
    if(ratio>bestRatio || (ratio===bestRatio && (CR[r[C]]||0)>(CR[bestCity]||0))){
      bestRatio=ratio; bestCity=r[C]; bestField=r[F]; bestDev=d;
    }
  }
  const kpis=[
    {label:'总数据点', value:rows.length, extra:''},
    {label:'整体一致率', value:a.rate.toFixed(1), unit:'%', extra:`一致 ${a.ok} / 有效 ${a.valid}`, warn:a.rate<40},
    {label:'覆盖城市', value:cities, extra:''},
    {label:'覆盖模块', value:mods, extra:''},
    {label:'最弱字段', value:weakest?weakest.f:'-', extra:weakest?`${modShow(weakest.m)} · ${weakest.rate.toFixed(1)}%`:'', warn:true},
    {label:'实况最大偏差', value:bestCity, cls:'kpi-name', extra:bestCity==='-'?'':`${bestField} 偏差${bestDev}（${bestRatio.toFixed(1)}倍）`, warn:bestRatio>=3},
  ];
  document.getElementById('kpis').innerHTML = kpis.map(k=>`
    <div class="kpi ${k.warn?'warn':''} ${k.cls||''}">
      <div class="label">${k.label}</div>
      <div class="value">${k.value}${k.unit?`<small>${k.unit}</small>`:''}</div>
      <div class="extra">${k.extra||''}</div>
    </div>`).join('');
}

// ====== 各图（统一用 G） ======
function drawModuleBar(){
  const rows = filterRows({...G, module:'全部'});
  const data = MODULE_LIST.map(m=>{
    const sub = rows.filter(r=>r[MO]===m);
    return {name:modShow(m), value:+aggRate(sub).rate.toFixed(1)};
  });
  EC.moduleBar.setOption({
    color:['#2563eb'], textStyle:BASE_TEXT, grid:{...GRID,left:44},
    tooltip:{trigger:'axis', formatter:p=>`${p[0].name}：${p[0].value}%`},
    xAxis:{type:'category',data:data.map(d=>d.name),axisLine:{lineStyle:{color:'#cbd5e1'}},axisLabel:{fontSize:13}},
    yAxis:{type:'value',max:100,axisLabel:{formatter:'{value}%',fontSize:12},splitLine:{lineStyle:{color:'#f1f5f9'}}},
    series:[{type:'bar',data:data.map(d=>({value:d.value,
      itemStyle:{color:d.value>=50?'#16a34a':d.value>=30?'#f59e0b':'#dc2626'}})),barWidth:'48%',
      label:{show:true,position:'top',formatter:'{c}%',fontSize:13,color:'#475569',fontWeight:600}}]
  },true);
}

let fieldBarAsc = true;   // true=一致率升序（低的在上）
let FB_MODULE = '全部';   // 字段一致率图独立模块筛选，覆盖全局
function toggleFieldSort(){ fieldBarAsc=!fieldBarAsc; drawFieldBar(); }
function setFbModule(v, btn){ FB_MODULE=v; document.querySelectorAll('#fb-module-list .sl-item').forEach(b=>b.classList.remove('active')); if(btn) btn.classList.add('active'); drawFieldBar(); }
function initFbModuleList(){
  const list = ['全部',...MODULE_LIST];
  document.getElementById('fb-module-list').innerHTML = '<div class="sl-title">模块筛选</div>' +
    list.map(m=>`<div class="sl-item${m===FB_MODULE?' active':''}" onclick="setFbModule('${m}',this)"><span>${m==='AQI模块'?'AQI':m}</span></div>`).join('');
}
function drawFieldBar(){
  const rows = filterRows({...G, field:'全部', module:FB_MODULE});
  const grp={};
  for(const r of rows){ const k=r[F]+'@'+r[MO]; if(!grp[k])grp[k]={f:r[F],m:r[MO],rows:[]}; grp[k].rows.push(r); }
  let arr = Object.values(grp).map(g=>({name:g.f+' ('+modShow(g.m)+')', value:+aggRate(g.rows).rate.toFixed(1)}));
  arr.sort((a,b)=> fieldBarAsc ? a.value-b.value : b.value-a.value);
  EC.fieldBar.setOption({
    textStyle:BASE_TEXT, grid:{left:170,right:50,top:20,bottom:20},
    tooltip:{trigger:'axis',formatter:p=>`${p[0].name}<br/>一致率 ${p[0].value}%`},
    xAxis:{type:'value',max:100,axisLabel:{formatter:'{value}%',fontSize:12},splitLine:{lineStyle:{color:'#f1f5f9'}}},
    yAxis:{type:'category',data:arr.map(d=>d.name),axisLine:{lineStyle:{color:'#cbd5e1'}},axisLabel:{fontSize:12}},
    series:[{type:'bar',data:arr.map(d=>({value:d.value,
      itemStyle:{color:d.value>=50?'#16a34a':d.value>=30?'#f59e0b':'#dc2626'}})),barWidth:'62%',
      label:{show:true,position:'right',formatter:'{c}%',fontSize:12,color:'#475569',fontWeight:600}}]
  },true);
  document.getElementById('fb-sort').textContent = fieldBarAsc ? '当前升序 ⇅ 切换' : '当前降序 ⇅ 切换';
  document.getElementById('fb-sort').classList.toggle('active', !fieldBarAsc);
  if(arr.length){
    const weakest = arr.reduce((a,b)=>a.value<b.value?a:b);
    const strongest = arr.reduce((a,b)=>a.value>b.value?a:b);
    const lowCnt = arr.filter(d=>d.value<30).length;
    document.getElementById('fieldBar-note').innerHTML =
      `共 ${arr.length} 个字段×模块组合；最高 <b>${strongest.name} ${strongest.value}%</b>，最低 <b>${weakest.name} ${weakest.value}%</b>${lowCnt?`，其中 <b>${lowCnt}</b> 个一致率不足 30% 需重点排查`:''}。`;
  } else {
    document.getElementById('fieldBar-note').innerHTML = '当前筛选下无数据。';
  }
}

function drawPie(){
  const rows = filterRows(G);
  const a = aggRate(rows);
  EC.pie.setOption({
    textStyle:BASE_TEXT, legend:{bottom:0,icon:'circle',textStyle:{fontSize:13}},
    tooltip:{trigger:'item',formatter:'{b}: {c} ({d}%)'},
    series:[{type:'pie',radius:['42%','68%'],center:['50%','45%'],
      itemStyle:{borderColor:'#fff',borderWidth:2},
      label:{formatter:'{b}\n{d}%',fontSize:13},
      data:[
        {name:'一致',value:a.ok,itemStyle:{color:'#16a34a'}},
        {name:'不一致',value:a.bad,itemStyle:{color:'#dc2626'}},
        {name:'缺数据',value:a.miss,itemStyle:{color:'#cbd5e1'}},
      ]}]
  },true);
}

const TREND24_CELLS = [];
function drawTrend24(){
  const rows = filterRows({...G, module:'24小时', period:'全部', field:'全部'});
  const fields = FIELD_ORDER.filter(f=>rows.some(r=>r[F]===f));
  const grid = document.getElementById('trend24-grid');
  // 字段数变化时重建 cell 实例
  if(TREND24_CELLS.length !== fields.length){
    TREND24_CELLS.forEach(c=>c.dispose());
    TREND24_CELLS.length = 0;
    grid.innerHTML = fields.map((f,i)=>`<div class="trend24-cell"><div class="trend24-title">${f.replace('天气现象','天气')}</div><div class="trend24-chart" id="trend24-cell-${i}"></div></div>`).join('');
    fields.forEach((f,i)=> TREND24_CELLS.push(echarts.init(document.getElementById('trend24-cell-'+i))) );
  } else {
    fields.forEach((f,i)=>{ const t=document.querySelectorAll('.trend24-title')[i]; if(t) t.textContent=f.replace('天气现象','天气'); });
  }
  fields.forEach((f,i)=>{
    const data = PERIOD_LIST.map(p=> +aggRate(rows.filter(r=>r[PR]===p && r[F]===f)).rate.toFixed(1) );
    TREND24_CELLS[i].setOption({
      textStyle:BASE_TEXT,
      grid:{left:30,right:10,top:22,bottom:22},
      tooltip:{trigger:'axis',formatter:p=>`${PERIOD_LIST[p[0].dataIndex].replace('时效','')}：${p[0].value}%`},
      xAxis:{type:'category',data:['短','中','长'],axisLine:{lineStyle:{color:'#cbd5e1'}},axisLabel:{fontSize:10},axisTick:{show:false}},
      yAxis:{type:'value',max:100,axisLabel:{formatter:'{value}%',fontSize:9},splitLine:{lineStyle:{color:'#f1f5f9'}},axisLine:{show:false},axisTick:{show:false}},
      series:[{type:'bar',data:data.map((v,j)=>({value:v,itemStyle:{color:['#2563eb','#60a5fa','#93c5fd'][j]}})),barWidth:'46%',
        label:{show:true,position:'top',formatter:'{c}',fontSize:10,color:'#475569',fontWeight:600}}]
    },true);
  });
}

let CITY_VIEW = 25;   // 城市排行显示条数：5/10/15/20/25/0(全部)
function setCityView(n, btn){
  CITY_VIEW = n;
  document.querySelectorAll('#city-view-ctrl .ctl-btn').forEach(b=>b.classList.remove('active'));
  if(btn) btn.classList.add('active');
  drawCityRank();
}
function drawCityRank(){
  const rows = filterRows({...G, city:'全部', cities:G.cities});
  const grp={};
  for(const r of rows){ if(!grp[r[C]])grp[r[C]]=[]; grp[r[C]].push(r); }
  let arr = Object.entries(grp).map(([c,rs])=>({name:c,value:+aggRate(rs).rate.toFixed(1)})).sort((a,b)=>a.value-b.value);
  // 右速览：最差5城 / 最好5城
  const worst = arr.slice(0,5);
  const best = arr.slice(-5).reverse();
  document.getElementById('city-rank-list').innerHTML = arr.length ? `
    <div class="sl-title">最差 5 城</div>
    ${worst.map(d=>`<div class="sl-item bad"><span>${d.name}</span><b>${d.value}%</b></div>`).join('')}
    <div class="sl-title">最好 5 城</div>
    ${best.map(d=>`<div class="sl-item ok"><span>${d.name}</span><b>${d.value}%</b></div>`).join('')}
  ` : '<div class="sl-empty">暂无数据</div>';
  EC.cityRank.setOption({
    textStyle:BASE_TEXT, grid:{left:100,right:70,top:20,bottom:20},
    dataZoom:[
      {type:'slider',yAxisIndex:0,right:10,width:10,
        start: CITY_VIEW===0 ? 0 : Math.max(0, Math.floor((arr.length-CITY_VIEW)/arr.length*100)),
        end: 100,
        backgroundColor:'transparent',
        dataBackground:{lineStyle:{color:'transparent'},areaStyle:{color:'transparent'}},
        selectedDataBackground:{lineStyle:{color:'#2563eb'},areaStyle:{color:'rgba(37,99,235,.18)'}},
        fillerColor:'rgba(37,99,235,.12)',borderColor:'transparent',
        handleSize:'100%',handleStyle:{color:'#2563eb',borderColor:'#1d4ed8',borderWidth:0},
        showDataShadow:false,showDetail:false}
    ],
    tooltip:{trigger:'axis',formatter:p=>`${p[0].name}：${p[0].value}%`},
    xAxis:{type:'value',max:100,axisLabel:{formatter:'{value}%',fontSize:12},splitLine:{lineStyle:{color:'#f1f5f9'}}},
    yAxis:{type:'category',data:arr.map(d=>d.name),axisLabel:{fontSize:11},axisLine:{lineStyle:{color:'#cbd5e1'}}},
    series:[{type:'bar',data:arr.map(d=>({value:d.value,
      itemStyle:{color:d.value>=50?'#16a34a':d.value>=30?'#f59e0b':'#dc2626'}})),barWidth:'60%',
      label:{show:true,position:'right',formatter:'{c}%',fontSize:10,color:'#64748b'}}]
  },true);
}

let DEV_FIELD = '温度';
function setDevField(f){ DEV_FIELD=f; drawDevHist(); }
function initDevFieldList(){
  const fields = FIELD_ORDER.filter(f=>!isWeather(f));
  if(!fields.includes(DEV_FIELD)) DEV_FIELD = fields[0] || '温度';
  document.getElementById('dev-field-list').innerHTML = '<div class="sl-title">字段（点击切换）</div>' +
    fields.map(f=>`<div class="sl-item${f===DEV_FIELD?' active':''}" onclick="setDevField('${f}')"><span>${f}</span></div>`).join('');
}
function drawDevHist(){
  let rows = filterRows({...G, field:'全部'}).filter(r=>r[F]===DEV_FIELD && num(r[DF])!==null);
  // 同步右栏高亮
  document.querySelectorAll('#dev-field-list .sl-item').forEach(el=>{
    const sp = el.querySelector('span'); el.classList.toggle('active', sp && sp.textContent===DEV_FIELD);
  });
  if(rows.length===0){ EC.devHist.clear(); return; }
  const diffs = rows.map(r=>num(r[DF]));
  const rawMax = Math.max(...diffs.map(Math.abs));
  const maxAbs = Math.max(5, Math.ceil(rawMax/2)*2);   // 量级按本字段，不再被气压长尾带偏
  const bins=21, step=(2*maxAbs)/bins, zeroIdx=Math.round(maxAbs/step);
  const counts=new Array(bins).fill(0);
  for(const d of diffs){ let idx=Math.floor((d+maxAbs)/step); if(idx<0)idx=0; if(idx>=bins)idx=bins-1; counts[idx]++; }
  const cats=counts.map((_,i)=>(-maxAbs+i*step).toFixed(0));
  EC.devHist.setOption({
    textStyle:BASE_TEXT, color:['#2563eb'], grid:GRID,
    title:{text:`${DEV_FIELD} 偏差分布（${rows.length} 样本，极差 ±${maxAbs}）`,left:'center',top:4,textStyle:{fontSize:12,color:'#475569',fontWeight:600}},
    tooltip:{trigger:'axis',formatter:p=>`偏差 ${p[0].name} 附近：${p[0].value} 条`},
    xAxis:{type:'category',data:cats,name:'差异',nameTextStyle:{fontSize:11},axisLine:{lineStyle:{color:'#cbd5e1'}},axisLabel:{fontSize:10}},
    yAxis:{type:'value',splitLine:{lineStyle:{color:'#f1f5f9'}},axisLabel:{fontSize:12}},
    series:[{type:'bar',data:counts,barWidth:'90%',
      itemStyle:{color:p=>{const v=Math.abs(+cats[p.dataIndex]); const g=maxAbs*0.1, y=maxAbs*0.3; return v<=g?'#16a34a':v<=y?'#f59e0b':'#dc2626';}},
      markLine:{silent:true,symbol:'none',data:[{xAxis:zeroIdx}],lineStyle:{color:'#dc2626',type:'dashed'},label:{formatter:'0',fontSize:11}}}]
  },true);
}

function drawSankey(){
  const rows = filterRows({...G, field:'全部'}).filter(r=>isWeather(r[F]) && r[CN] && r[IV]);
  const pairs={};
  for(const r of rows){ const k=r[CN]+'||'+r[IV]; pairs[k]=(pairs[k]||0)+1; }
  // 只取不一致配对 top15
  const sorted = Object.entries(pairs).filter(([k])=>{const [a,b]=k.split('||');return a!==b;})
    .sort((a,b)=>b[1]-a[1]).slice(0,15);
  const nodeset=new Set(), links=[];
  for(const [k,v] of sorted){
    const [a,b]=k.split('||');
    const la='国内·'+a, lb='海外·'+b;
    nodeset.add(la); nodeset.add(lb);
    links.push({source:la, target:lb, value:v});
  }
  const nodes=[...nodeset].map(n=>({name:n}));
  // 解读条
  const noteEl = document.getElementById('sankey-note');
  if(sorted.length===0){
    noteEl.innerHTML = '当前筛选下无不一致项，国内外天气现象判断完全一致。';
  } else {
    const [a,b]=sorted[0][0].split('||');
    noteEl.innerHTML = `仅展示不一致项 TOP15 流向，条带越粗表示该误判越频繁。最常见误判：<b>国内${a} › 海外${b}</b>（${sorted[0][1]}次）。`;
  }
  // 流向统计表（翻译桑基图为表格：国内天气 | 海外天气 | 次数）
  const tblEl = document.getElementById('sankey-table');
  if(sorted.length===0){
    tblEl.innerHTML = '';
  } else {
    const top10 = sorted.slice(0,10);
    const left = top10.slice(0,5), right = top10.slice(5);
    const mk = (arr,start)=> arr.length ? `<table><thead><tr><th class="rank">#</th><th>国内天气</th><th>海外天气</th><th>次数</th></tr></thead><tbody>` +
      arr.map((x,i)=>{const [a,b]=x[0].split('||');return `<tr><td class="rank">${start+i}</td><td>${a}</td><td>${b}</td><td><b>${x[1]}</b></td></tr>`;}).join('') + `</tbody></table>` : '';
    tblEl.innerHTML = `<div class="sankey-twin"><div class="sankey-col"><div class="sl-title">前 5</div>${mk(left,1)}</div><div class="sankey-col"><div class="sl-title">6 - 10</div>${mk(right,6)}</div></div>`;
  }
  EC.sankey.setOption({
    textStyle:BASE_TEXT,
    tooltip:{formatter:p=>p.dataType==='edge'?`${p.data.source} → ${p.data.target}：${p.data.value}次`:`${p.name}`},
    series:[{type:'sankey', data:nodes, links,
      nodeWidth:20, nodeGap:14, layout:'none',
      lineStyle:{color:'gradient', opacity:.65, curveness:.5},
      itemStyle:{borderWidth:0,borderColor:'#fff'},
      label:{fontSize:12, color:'#334155'},
      emphasis:{focus:'adjacency',lineStyle:{opacity:.88}},
      left:8, right:90, top:20, bottom:20
    }]
  },true);
}

// ====== 分模块 TOP5 表 ======
const TOP5_DEFS = [
  {title:'实况模块 · TOP5 偏差城市', module:'实况', period:'全部'},
  {title:'24小时 · 短时效(1-6h) TOP5 偏差城市', module:'24小时', period:'短时效(1-6h)'},
  {title:'24小时 · 中时效(7-12h) TOP5 偏差城市', module:'24小时', period:'中时效(7-12h)'},
  {title:'24小时 · 长时效(13-24h) TOP5 偏差城市', module:'24小时', period:'长时效(13-24h)'},
  {title:'15天预报 · TOP5 偏差城市', module:'15天', period:'全部'},
  {title:'AQI模块 · TOP5 偏差城市', module:'AQI模块', period:'全部'},
];
function drawModuleTop5(def, elId){
  const rows = filterRows({...G, module:def.module, field:'全部', period:def.period});
  const fields = FIELD_ORDER.filter(f=>rows.some(r=>r[F]===f));
  const trs = fields.map(f=>{
    const sub = rows.filter(r=>r[F]===f);
    // 按城市去重：同一城市只保留偏差最大的一条（与xlsx「前五偏差城市」口径一致），避免同城市多时次重复上榜
    const cityBest = {};
    for(const r of sub){
      const d = num(r[DF]);
      if(d===null) continue;
      const score = isWeather(f) ? d : Math.abs(d);
      if(!(r[C] in cityBest) || score > cityBest[r[C]].score){
        cityBest[r[C]] = {r, score};
      }
    }
    const WL=DATA.weatherTexts||{};
    const wld=(cn,iv)=>{const a=WL[cn],b=WL[iv]; if(!a||!b) return 0; return Math.abs(a.level-b.level);};
    const sorted = Object.values(cityBest).sort((a,b)=>{
      const dd=b.score-a.score;
      if(dd) return dd;
      return wld(b.r[CN],b.r[IV]) - wld(a.r[CN],a.r[IV]);
    }).slice(0,5);
    const tds = sorted.map((item, idx)=>{
      const r = item.r;
      const d = num(r[DF]);
      const topClass = 'top'+(idx+1);   // top1~top5 五档配色
      if(isWeather(f)){
        return `<td class="${topClass}">${r[C]}<br><small class="${devClass(d)}">${r[CN]}→${r[IV]}</small></td>`;
      }
      return `<td class="${topClass}">${r[C]}<br><small class="${devClass(d)}">${d}</small></td>`;
    });
    while(tds.length<5) tds.push('<td><small>-</small></td>');
    return `<tr><td class="fname">${f}</td>${tds.join('')}</tr>`;
  });
  document.getElementById(elId).innerHTML = `
    <div class="tbl-wrap"><table>
      <thead><tr><th>字段</th><th>TOP1</th><th>TOP2</th><th>TOP3</th><th>TOP4</th><th>TOP5</th></tr></thead>
      <tbody>${trs.join('')}</tbody>
    </table></div>`;
}
function drawAllTop5(){
  let defs = TOP5_DEFS;
  if(G.module!=='全部') defs = defs.filter(d=>d.module===G.module);
  const zone = document.getElementById('top5-zone');
  zone.innerHTML = defs.map((d,i)=>`
    <div class="card span2" style="margin-bottom:16px">
      <div class="card-head"><span class="card-title">${d.title}</span></div>
      <div id="top5-${i}"></div>
    </div>`).join('');
  defs.forEach((d,i)=>drawModuleTop5(d, 'top5-'+i));
}

// ====== 明细表（按需查询，独立于全局筛选，避免一次性渲染全量）======
const DQ = {module:'全部', field:'全部', period:'全部', search:'', loaded:false};
const DPAGE = {page:1, size:50, sortKey:C, asc:true, search:''};
function detailRowsForQuery(){
  // 主查询条件：明细表自己的模块/字段/时效 + 全局地区；不随全局模块/字段/时效联动
  let rows = filterRows({module:DQ.module, field:DQ.field, period:DQ.period, cities:G.cities});
  const s = DPAGE.search.trim();
  if(s) rows = rows.filter(r=>r[C].indexOf(s)>=0 || r[F].indexOf(s)>=0 || r[MO].indexOf(s)>=0);
  return rows;
}
function drawDetail(){
  let rows = detailRowsForQuery();
  const sk=DPAGE.sortKey, asc=DPAGE.asc;
  rows = rows.slice().sort((a,b)=>{
    let x=a[sk], y=b[sk];
    if(sk===CN||sk===IV||sk===DF){ x=num(x); y=num(y); if(x===null)x=-Infinity; if(y===null)y=-Infinity; }
    else { x=''+x; y=''+y; }
    if(x<y) return asc?-1:1; if(x>y) return asc?1:-1; return 0;
  });
  const total=rows.length;
  const pages=Math.max(1,Math.ceil(total/DPAGE.size));
  if(DPAGE.page>pages) DPAGE.page=pages;
  const start=(DPAGE.page-1)*DPAGE.size;
  const page=rows.slice(start,start+DPAGE.size);
  const cols=[['城市',C],['模块',MO],['字段',F],['时次',TS],['国内值',CN],['海外值',IV],['差异',DF],['是否一致',OK],['时效',PR]];
  const th=cols.map(([n,i])=>`<th class="sortable" onclick="sortDetail(${i})">${n}${DPAGE.sortKey===i?(DPAGE.asc?' ↑':' ↓'):''}</th>`).join('');
  const tb=page.map(r=>{
    const tag = r[OK]==='一致'?'<span class="tag ok">一致</span>':r[OK]==='不一致'?'<span class="tag bad">不一致</span>':'<span class="tag miss">缺数据</span>';
    return `<tr><td class="city">${r[C]}</td><td>${modShow(r[MO])}</td><td>${r[F]}</td><td class="num">${r[TS]}</td>
      <td class="num">${r[CN]||'-'}</td><td class="num">${r[IV]||'-'}</td><td class="num dev">${r[DF]||'-'}</td><td>${tag}</td><td>${r[PR]||'-'}</td></tr>`;
  }).join('');
  document.getElementById('ch-detail').innerHTML=`
    <div class="tbl-tools">
      <input placeholder="🔍 在结果中二次搜索城市/字段/模块..." value="${DPAGE.search}" oninput="searchDetail(this.value)">
      <span>共 ${total} 条</span>
      <div class="pager">
        <button onclick="pageDetail(-1)" ${DPAGE.page<=1?'disabled':''}>上一页</button>
        <span>${DPAGE.page} / ${pages}</span>
        <button onclick="pageDetail(1)" ${DPAGE.page>=pages?'disabled':''}>下一页</button>
      </div>
    </div>
    <table><thead><tr>${th}</tr></thead><tbody>${tb}</tbody></table>`;
}
function queryDetail(){
  const dq={module:document.getElementById('d-module').value, field:document.getElementById('d-field').value, period:document.getElementById('d-period').value};
  const search=document.getElementById('d-search').value.trim();
  const tip=document.getElementById('dq-tip');
  tip.classList.remove('err');
  // 至少一个非"全部"条件，避免全量加载
  const hasFilter = dq.module!=='全部' || dq.field!=='全部' || dq.period!=='全部' || search.length>0;
  if(!hasFilter){
    tip.classList.add('err');
    tip.innerHTML='⚠ 请至少选择一个筛选维度或输入搜索词，避免全量加载';
    return;
  }
  Object.assign(DQ, dq, {search, loaded:true});
  DPAGE.search=''; DPAGE.page=1;
  const rows = detailRowsForQuery();
  if(rows.length>5000){
    tip.classList.add('err');
    tip.innerHTML=`⚠ 匹配 ${rows.length} 条，超过单次上限 5000 条，请增加筛选条件缩小范围`;
    DQ.loaded=false;
    document.getElementById('detail-result').style.display='none';
    document.getElementById('d-collapse').style.display='none';
    return;
  }
  tip.innerHTML=`✅ 匹配 ${rows.length} 条（全量约 ${detail.length} 条）`;
  document.getElementById('detail-result').style.display='';
  document.getElementById('d-collapse').style.display='';
  drawDetail();
}
function collapseDetail(){
  DQ.loaded=false;
  document.getElementById('detail-result').style.display='none';
  document.getElementById('d-collapse').style.display='none';
  const tip=document.getElementById('dq-tip');
  tip.classList.remove('err');
  tip.innerHTML=`提示：至少选择一个筛选维度或输入搜索词，避免全量加载。全量约 ${detail.length} 条`;
}
function initDetail(){
  const fill=(id,list)=>{ document.getElementById(id).innerHTML = list.map(o=>`<option value="${o}">${o==='AQI模块'?'AQI':o}</option>`).join(''); };
  fill('d-module', ['全部',...MODULE_LIST]);
  fill('d-field', ['全部',...FIELD_LIST]);
  fill('d-period', ['全部',...PERIOD_LIST]);
  document.getElementById('dq-total').textContent = detail.length;
}
function sortDetail(i){ if(DPAGE.sortKey===i)DPAGE.asc=!DPAGE.asc; else{DPAGE.sortKey=i;DPAGE.asc=true;} DPAGE.page=1; drawDetail(); }
function searchDetail(v){ DPAGE.search=v; DPAGE.page=1; drawDetail(); }
function pageDetail(d){ DPAGE.page+=d; drawDetail(); }

// ====== 全局应用 ======
function applyVisibility(){
  const rows = filterRows(G);
  const show=(id,ok)=>{ const el=document.getElementById(id); if(el) el.classList.toggle('hidden', !ok); };
  // 选地区后不适配的表隐藏：单城市时城市排行无意义；各图无数据则隐藏
  show('card-cityRank', G.cities.length!==1 && rows.length>0);
  show('card-moduleBar', filterRows({...G,module:'全部'}).length>0);
  show('card-pie', rows.length>0);
  show('card-fieldBar', filterRows({...G,field:'全部'}).length>0);
  show('card-trend24', filterRows({...G,module:'24小时',period:'全部',field:'全部'}).length>0);
  show('card-devHist', filterRows({...G,field:'全部'}).some(r=>!isWeather(r[F]) && num(r[DF])!==null));
  show('card-sankey', filterRows({...G,field:'全部'}).some(r=>isWeather(r[F]) && r[CN] && r[IV]));
}
function applyGlobal(){
  G.module = document.getElementById('g-module').value;
  G.field = document.getElementById('g-field').value;
  G.period = document.getElementById('g-period').value;
  G.cities = regionCities();
  applyVisibility();
  drawKPI();
  drawModuleBar(); drawFieldBar(); drawPie(); drawTrend24();
  drawCityRank(); drawDevHist(); drawSankey();
  drawAllTop5();
  // 明细表独立按需加载：仅在已展开时随地区刷新
  if(DQ.loaded) drawDetail();
  // 显隐变化后重算图表尺寸
  requestAnimationFrame(()=>{ for(const id in EC){ if(EC[id]) EC[id].resize(); } TREND24_CELLS.forEach(c=>c.resize()); });
}
function resetGlobal(){
  document.getElementById('g-module').value='全部';
  document.getElementById('g-field').value='全部';
  document.getElementById('g-period').value='全部';
  initRegion();
  applyGlobal();
}

// ====== 地区三级联动 ======
function regionCities(){
  const r1=document.getElementById('r1').value;
  const r2=document.getElementById('r2').value;
  const r3=document.getElementById('r3').value;
  if(!r1 || r1==='全部') return [];
  if(!r2) return Object.values(REGION_TREE[r1]).flat();
  if(!r3) return (REGION_TREE[r1][r2]||[]).slice();
  return [r3];
}
function initRegion(){
  const r1=document.getElementById('r1');
  r1.innerHTML = '<option value="全部">全部地区</option>' +
    Object.keys(REGION_TREE).map(c=>`<option value="${c}">${c}</option>`).join('');
  document.getElementById('r2').innerHTML='';
  document.getElementById('r3').innerHTML='';
}
function onRegion1(val){
  const r2=document.getElementById('r2'), r3=document.getElementById('r3');
  r3.innerHTML='';
  if(!val || val==='全部'){ r2.innerHTML=''; applyGlobal(); return; }
  r2.innerHTML = '<option value="">全部</option>' +
    Object.keys(REGION_TREE[val]).sort().map(p=>`<option value="${p}">${p}</option>`).join('');
  applyGlobal();
}
function onRegion2(val){
  const r3=document.getElementById('r3');
  const r1=document.getElementById('r1').value;
  if(!val){ r3.innerHTML=''; applyGlobal(); return; }
  r3.innerHTML = '<option value="">全部</option>' +
    (REGION_TREE[r1][val]||[]).map(c=>`<option value="${c}">${c}</option>`).join('');
  applyGlobal();
}
function onRegion3(val){ applyGlobal(); }

// ====== 筛选器初始化 ======
function initFilters(){
  const fill=(id,list)=>{ document.getElementById(id).innerHTML =
    list.map(o=>`<option value="${o}">${o==='AQI模块'?'AQI':o}</option>`).join(''); };
  fill('g-module', ['全部',...MODULE_LIST]);
  fill('g-field', ['全部',...FIELD_LIST]);
  fill('g-period', ['全部',...PERIOD_LIST]);
}

function initAll(){
  initFilters();
  initRegion();
  initDetail();
  initDevFieldList();
  initFbModuleList();
  initCharts();
  applyGlobal();
}
initAll();
</script>
</body>
</html>"""

    html = html.replace('__ECHARTS__', echarts_js)
    html = html.replace('__DATA__', data_json)
    html = html.replace('__TIME__', meta['time'])
    html = html.replace('__SOURCE__', meta['source'])
    html = html.replace('__KOUJING__', meta['koujing'])
    html = html.replace('__CITIES__', str(meta['cities']))
    html = html.replace('__POINTS__', str(meta['points']))
    html = html.replace('__TH_ROWS__', th_rows)
    rain_section = """    <h4>二、降水量等级比对</h4>
    <p>降水量不设数值阈值，采用等级比对：国内值/海外值分别映射到降水量等级，等级相同即判一致。</p>
    <table><thead><tr><th>等级</th><th>降水量区间 (mm)</th></tr></thead><tbody>__RAIN_ROWS__</tbody></table>""" if rain_th else ''
    html = html.replace('__RAIN_SECTION__', rain_section)
    html = html.replace('__RAIN_ROWS__', rain_rows)
    return html


# =========================================================
# 主流程
# =========================================================

def main():
    xlsx_arg = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx, csv_path = find_inputs(xlsx_arg)
    print(f"读取数据:")
    print(f"  xlsx: {os.path.basename(xlsx)}")
    print(f"  csv : {os.path.basename(csv_path) if csv_path else '无'}")

    summary = read_summary(xlsx)
    detail = read_detail(csv_path) if csv_path else []
    thresholds = read_thresholds()
    import yaml as _yml
    rain_th = _yml.safe_load(open(CONFIG_PATH, encoding='utf-8')).get('rain_thresholds', {}) or {}
    echarts_js = read_echarts()

    cities = len(set(r[0] for r in detail)) if detail else 0
    points = len(detail)

    base_xlsx = os.path.basename(xlsx)
    koujing = '阈值口径'
    if '均值' in base_xlsx:
        koujing = '阈值口径（6次均值）'
    try:
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        if '说明' in wb.sheetnames:
            ws = wb['说明']
            for row in ws.iter_rows(values_only=True):
                if row[0] and '比对口径' in str(row[0]):
                    koujing = str(row[0]).replace('比对口径: ', '').replace('比对口径:', '').strip()
                    break
    except Exception:
        pass

    meta = {
        'time': parse_window_end(xlsx),
        'source': base_xlsx,
        'koujing': koujing,
        'cities': cities,
        'points': points,
    }

    data = {
        'meta': meta,
        'summary': summary,
        'detail': detail,
        'region': CITY_REGION,
        'thresholds': thresholds,
        'weatherTexts': read_weather_texts(),
        'cityRank': load_city_rank(),
    }
    data_json = json.dumps(data, ensure_ascii=False)

    html = build_html(data_json, echarts_js, meta, thresholds, rain_th)

    out_path = os.path.join(OUT_DIR, f'一致性比对报告_HTML_{TIMESTAMP}.html')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)

    size_mb = os.path.getsize(out_path) / 1024 / 1024
    print(f"\n✅ HTML 报告: {out_path}")
    print(f"   体积: {size_mb:.2f} MB")
    print(f"   数据点: {points}，城市: {cities}")
    print(f"   图表: 7 个 ECharts 图 + 6 张分模块TOP5表 + 明细表")
    print(f"   交互: 全局筛选(模块/字段/时效/地区三级联动) + 字段条形升降序 + 明细表排序分页")
    print(f"\n用浏览器打开即可，已内联 echarts，离线可用。")


if __name__ == '__main__':
    main()
