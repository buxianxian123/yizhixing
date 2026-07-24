#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
一站式比对：拉取实时数据 → 严格相等比对 + 阈值口径比对
一次运行输出两份 xlsx 报告（带时间戳，永不覆盖）
"""
import json, csv, os, datetime, hashlib, hmac, subprocess, time
import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

# ====== 接口鉴权配置 ======
PASSWORD_CN = '49ff9a4e5e8bd5e8ce9e057c5adc5d2d'
PASSWORD_IN = '923ffbda8b65bf0f8e126824d050887a'
TOKEN_CN = 'cc920d85f8fbb762b6c705375add6c32'
TOKEN_IN = 'b88b7a5375e293671270016fe556a4b5'

# ====== 路径配置 ======
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.join(SCRIPT_DIR, '..', 'data')
CITY_CSV = os.path.join(BASE, '天气一致性测试城市_热门城市筛选.csv')
OUT_DIR = os.path.join(BASE, '比对结果')
TIMESTAMP = datetime.datetime.now().strftime('%Y%m%d_%H%M')
CITY_RANK = {}  # 城市序号，偏差相同时偏远地区优先
CSV_PATH = os.path.join(OUT_DIR, f'数据明细_{TIMESTAMP}.csv')
XLSX_STRICT = os.path.join(OUT_DIR, f'一致性比对报告_严格相等_{TIMESTAMP}.xlsx')
XLSX_THRESHOLD = os.path.join(OUT_DIR, f'一致性比对报告_阈值口径_{TIMESTAMP}.xlsx')
CONFIG_PATH = os.path.join(SCRIPT_DIR, 'compare_config.yaml')

# ====== 读取阈值配置 ======
config = yaml.safe_load(open(CONFIG_PATH, encoding='utf-8'))
THRESHOLDS = config['thresholds']
WIND_CFG = config['wind_convert']
MODULES = config['modules']
WEATHER_MAP = config.get('weather_mapping', {})
WTH_TEXTS = WEATHER_MAP.get('texts', {})

# ====== 降水量等级配置 ======
_rain_raw = config.get('rain_thresholds', {})
RAIN_TH = []          # [(上界, 等级名), ...]
RAIN_NAMES = []       # [等级名, ...]
for name, th in _rain_raw.items():
    if th == '~' or th is None:
        RAIN_TH.append((float('inf'), name))
    else:
        RAIN_TH.append((float(th), name))
    RAIN_NAMES.append(name)
# 按上界升序排序
RAIN_TH.sort(key=lambda x: x[0])

# ====== 24小时时效分段 ======
PERIODS_24H = [
    ('短时效(1-6h)', 0, 6),
    ('中时效(7-12h)', 6, 12),
    ('长时效(13-24h)', 12, 24),
]

# =========================================================
# 第一部分：API 请求
# =========================================================

def cn_key(lat, lon):
    return hashlib.md5((PASSWORD_CN + '0' + lat + lon).encode()).hexdigest()

def in_key(lat, lon):
    return hmac.new(PASSWORD_IN.encode(), ('0' + lat + lon).encode(), hashlib.sha256).hexdigest()

def cn_url(lat, lon):
    return f'http://coapi.moji.com/whapi/v2/weather?timestamp=0&language=zh-CN&token={TOKEN_CN}&lat={lat}&lon={lon}&key={cn_key(lat, lon)}'

def in_url_full(lat, lon):
    return f'https://datasw1.api.moweather.com/whapi/in/weather?token={TOKEN_IN}&lon={lon}&lat={lat}&lang=zh-CN&current=1&hourly=24&daily=15&aqi=1&metric=true&ts=0&key={in_key(lat, lon)}'

def fetch(url, retry=2):
    """请求接口，重试 retry 次"""
    for _ in range(retry + 1):
        r = subprocess.run(['curl', '-sk', '--max-time', '25', url], capture_output=True, text=True)
        try:
            d = json.loads(r.stdout)
            if d.get('code') == 0:
                return d.get('data', {})
        except:
            pass
        time.sleep(0.5)
    return None

def fetch_city(name, lon, lat):
    """拉单个城市的国内+国际数据，任一失败对应项为 None。供一次性脚本和定时脚本复用"""
    cn_data = fetch(cn_url(lat, lon))
    in_data = fetch(in_url_full(lat, lon))
    return cn_data, in_data

# =========================================================
# 第二部分：比对工具函数
# =========================================================

def num(v):
    if v is None: return None
    try: return float(v)
    except: return None

def wind_convert(v):
    if v is None or not WIND_CFG.get('enabled'): return v
    return round(v / WIND_CFG['divisor'], 2)

def text_diff(cv, iv):
    if cv is None or iv is None: return ''
    return '一致' if cv == iv else '不一致'

def rain_level(v):
    """降水量mm → 等级索引(从0开始), 等级定义来自 compare_config.yaml rain_thresholds"""
    if v is None: return None
    for i, (th, _) in enumerate(RAIN_TH):
        if v < th:
            return i
    return len(RAIN_TH) - 1

def get_threshold(field):
    """按字段名匹配阈值（支持'温度(最高)'等带括号字段）"""
    for k, v in THRESHOLDS.items():
        if k in field: return v
    return None

def get_period_label(source, idx):
    if source != 'hourly': return ''
    for label, start, end in PERIODS_24H:
        if start <= idx < end: return label
    return ''

def calc_weather_deviation(cn_text, intl_text):
    """
    天气现象语义映射偏差计算（五分制评分 → 偏差值）
    将中文天气文字映射到(大类,量级,高影响标记)，按评分规则打分

    评分规则（5=最佳 → 0=最差，仅5分算一致）：
      5分  主天气一致+量级一致                     → 完全匹配
      4分  主天气一致+量级差1级                     → 轻微量级偏差
      3分  主天气不一致（晴↔多云等），均非高影响      → 主天气不一致
      2分  主天气错判（量级差≥2），或降水vs非降水    → 明显偏差
      1分  涉及高影响天气+同大类                   → 高影响偏差
      0分  高影响天气+不同大类                     → 高影响漏报/错判

    返回 (偏差, '一致'/'不一致')
    """
    if cn_text is None or intl_text is None:
        return (None, '')

    a = WTH_TEXTS.get(cn_text)
    b = WTH_TEXTS.get(intl_text)
    ok_min = WEATHER_MAP.get('ok_min_score', 5)

    # 未识别的天气文字 → 回退到文字比对
    if a is None or b is None:
        ok = '一致' if str(cn_text) == str(intl_text) else '不一致'
        score = ok_min if ok == '一致' else 3
        deviation = ok_min - score
        return (deviation, ok)

    level_diff = abs(a['level'] - b['level'])

    if a['cat'] == b['cat']:
        if level_diff == 0:
            score = WEATHER_MAP['score_same_cat_same_level']
        elif level_diff == 1:
            score = WEATHER_MAP['score_same_cat_level_diff_1']
        else:
            # level_diff >= 2
            if a['hi'] or b['hi']:
                score = WEATHER_MAP['score_hi_same_cat']
            else:
                score = WEATHER_MAP['score_same_cat_level_diff_ge2']
    else:
        if a['hi'] or b['hi']:
            score = WEATHER_MAP['score_hi_diff_cat']
        else:
            # 一方雨/雪一方非降水 → 比纯云量差异更严重
            is_precip = lambda t: t['cat'] in ('雨', '雪')
            if is_precip(a) != is_precip(b):
                score = WEATHER_MAP.get('score_diff_cat_precip', 2)
            else:
                score = WEATHER_MAP['score_diff_cat_no_hi']

    deviation = ok_min - score
    ok = '一致' if score >= ok_min else '不一致'
    return (deviation, ok)

def cmp_point(city, module, field, ts, cnv, iv, spec, period='', strict=False):
    """
    单个数据点比对
    strict=True  → 严格相等（diff==0）
    strict=False → 阈值判断（|diff|<=threshold）
    返回 10 元组
    """
    typ = spec.get('type', 'numeric'); note = spec.get('note', '')
    if typ == 'wind':
        cv = num(cnv); iv_conv = wind_convert(num(iv))
        if WIND_CFG.get('enabled') and iv is not None:
            note = f"海外原{iv} ÷{WIND_CFG['divisor']}"
    elif typ == 'weather':
        cv = cnv; iv_conv = iv
    else:
        cv = num(cnv); iv_conv = num(iv)

    if typ == 'weather':
        diff, ok = calc_weather_deviation(cnv, iv)
        return (city, module, field, ts, cnv, iv, diff, ok, '按语义映射比对', period)

    if typ == 'rain_level':
        cv = num(cnv); iv_conv = num(iv)
        if cv is None or iv_conv is None:
            return (city, module, field, ts, cnv, iv, '', '缺数据', note, period)
        cl = rain_level(cv); il = rain_level(iv_conv)
        diff = round(abs(cv - iv_conv), 2)
        ok = '一致' if cl == il else '不一致'
        cn_lv = RAIN_NAMES[cl] if cl is not None else '?'
        in_lv = RAIN_NAMES[il] if il is not None else '?'
        return (city, module, field, ts, cnv, iv, diff, ok, f'国内{cn_lv} vs 海外{in_lv}', period)

    if cv is None or iv_conv is None:
        return (city, module, field, ts, cnv, iv, '', '缺数据', note, period)

    diff = round(abs(cv - iv_conv), 2)
    if strict:
        ok = '一致' if diff == 0 else '不一致'
    else:
        th = get_threshold(field)
        ok = '一致' if (th is not None and abs(diff) <= th) else '不一致'
    return (city, module, field, ts, cv, iv_conv, diff, ok, note, period)

def build_points(city, cn, ind, strict=False):
    """遍历所有模块/字段，返回该城市比对数据点列表"""
    P = []
    for module, mspec in MODULES.items():
        source = mspec['source']; fields = mspec['fields']
        if mspec.get('multi'):
            cn_arr = cn.get(source, [])[:mspec.get('limit', 99)]
            ind_arr = ind.get(source, [])[:mspec.get('limit', 99)]
            ts_key = mspec.get('ts_key'); lim = mspec.get('limit', 99)
            for k in range(min(len(cn_arr), len(ind_arr), lim)):
                a = cn_arr[k]; b = ind_arr[k]
                ts = a.get(ts_key, f'第{k+1}') if ts_key else f'第{k+1}'
                period = get_period_label(source, k)
                for field, spec in fields.items():
                    P.append(cmp_point(city, module, field, ts, a.get(spec['cn']), b.get(spec['intl']), spec, period, strict))
        else:
            cn_mod = cn.get(source, {}); ind_mod = ind.get(source, {})
            ts = mspec.get('ts_label', '')
            for field, spec in fields.items():
                P.append(cmp_point(city, module, field, ts, cn_mod.get(spec['cn']), ind_mod.get(spec['intl']), spec, '', strict))
    return P

# =========================================================
# 第三部分：生成 xlsx
# =========================================================

def _weather_level_diff(cn_val, intl_val):
    """天气现象误判的量级差(并列时细分严重度:特大暴雨>暴雨>大雨),非天气返回0"""
    a = WTH_TEXTS.get(cn_val); b = WTH_TEXTS.get(intl_val)
    if not a or not b: return 0
    return abs(a['level'] - b['level'])

def gen_xlsx(allP, title, xlsx_path, extra_notes=None):
    """生成一致性比对报告 xlsx
    extra_notes: 可选，追加到「说明」sheet 末尾的额外说明（如均值报告的采样信息），默认 None 不追加"""
    # 从title解析均值次数(如"阈值口径(47次均值)"),天气误判次数标注"X次/N份"避免歧义
    import re
    _m = re.search(r'(\d+)次均值', title)
    _avg_n = int(_m.group(1)) if _m else 0
    _cities = len(set(p[0] for p in allP))
    def _cnt(n): return f'{n}/{_cities*_avg_n}' if _avg_n > 1 else f'{n}次'

    # ---------- Sheet1 数据明细 ----------
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '数据明细'
    H = ['城市', '模块', '字段', '时次', '国内值', '海外值', '差异', '是否一致', '备注', '时效分段']
    ws.append(H)
    for c in range(1, len(H)+1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    green = PatternFill('solid', fgColor='E6F7E6')
    red = PatternFill('solid', fgColor='FFE6E6')
    gray = PatternFill('solid', fgColor='F0F0F0')
    for p in allP:
        ws.append(list(p))
    for i, p in enumerate(allP, 2):
        cell = ws.cell(row=i, column=8)
        if p[7] == '一致': cell.fill = green
        elif p[7] == '不一致': cell.fill = red
        else: cell.fill = gray
        # 差异列(column 7) 数值带单位
        diff_cell = ws.cell(row=i, column=7)
        if isinstance(p[6], (int, float)):
            f = p[2]
            u = ''
            if '天气现象' not in f:
                if '温度' in f or '体感' in f: u = '℃'
                elif '湿度' in f: u = '%'
                elif '风速' in f: u = 'm/s'
                elif '气压' in f: u = 'hPa'
                elif '降水' in f: u = 'mm'
            if u:
                diff_cell.value = f'{p[6]}{u}'
    for col, w in zip('ABCDEFGHIJ', [12, 10, 14, 18, 14, 14, 10, 10, 22, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'

    # ---------- Sheet2 总结 ----------
    ws2 = wb.create_sheet('总结')
    stat = defaultdict(lambda: {'total': 0, 'miss': 0, 'n': 0, 'ok': 0, 'sumdiff': 0, 'maxdiff': 0, 'maxcity': '', 'top': {}, 'dev_counts': defaultdict(int), 'pair_counts': defaultdict(int)})
    for p in allP:
        city, module, field, ts, cnv, iv, diff, ok, note, period = p
        s = stat[(field, module, period)]
        s['total'] += 1
        if ok in ('缺数据', ''):
            s['miss'] += 1; continue
        s['n'] += 1
        if ok == '一致': s['ok'] += 1
        if isinstance(diff, (int, float)):
            s['sumdiff'] += abs(diff)
            # 最大偏差城市: 偏差更大则替换; 偏差相同(并列)取更偏远的城市
            # (和「前五偏差城市」口径一致, 避免并列时被 CSV 顺序靠前的北京占位)
            if abs(diff) > abs(s['maxdiff']) or \
               (abs(diff) == abs(s['maxdiff']) and CITY_RANK.get(city, 9999) > CITY_RANK.get(s['maxcity'], 9999)):
                s['maxdiff'] = diff; s['maxcity'] = city
            if city not in s['top'] or abs(diff) > abs(s['top'][city][0]):
                s['top'][city] = (diff, str(cnv or ''), str(iv or ''))
            s['dev_counts'][int(diff)] += 1
        # 天气现象字段：统计 CN→INTL 配对频次
        if '天气现象' in field and cnv is not None and iv is not None:
            s['pair_counts'][(str(cnv), str(iv))] += 1

    H2 = ['字段', '模块', '时效', '总数据', '缺数据(已排除)', '有效样本', '一致数', '一致率', '平均偏差', '最大偏差', '最大偏差城市']
    ws2.append(H2)
    for c in range(1, len(H2)+1):
        ws2.cell(row=1, column=c).font = Font(bold=True)
    PERIOD_ORDER = {p[0]: i for i, p in enumerate(PERIODS_24H)}
    for module in MODULES.keys():
        items = sorted(stat.items(), key=lambda x: (
            list(MODULES.keys()).index(x[0][1]) if x[0][1] in MODULES else 99,
            x[0][0],
            PERIOD_ORDER.get(x[0][2], 99)
        ))
        for (field, m, period), s in items:
            if m != module: continue
            rate = f"{s['ok']/s['n']*100:.1f}%" if s['n'] else '0'

            # 天气现象字段：显示最频繁的 CN→INTL 误判对
            if '天气现象' in field and s['pair_counts']:
                sorted_pairs = sorted(s['pair_counts'].items(), key=lambda x: -x[1])
                mismatch = [(p, c) for p, c in sorted_pairs if p[0] != p[1]]  # 排除一致的
                # 平均偏差：最常见的误判对，只写一个
                if mismatch:
                    avg_display = f'国内{mismatch[0][0][0]}→国外{mismatch[0][0][1]}({_cnt(mismatch[0][1])})'
                else:
                    avg_display = round(s['sumdiff'] / s['n'], 2) if s['n'] else ''
                # 最大偏差：按实际偏差等级算，找最高等级的常见误判对
                max_dev = max(s['dev_counts'].keys(), key=lambda k: abs(k)) if s['dev_counts'] else None
                if max_dev and max_dev >= 1:
                    # 逐一算每个误判对的偏差值，找到等于 max_dev 的
                    best_pair = None
                    best_cnt = 0
                    for (cn_val, intl_val), cnt in mismatch:
                        dev, _ = calc_weather_deviation(cn_val, intl_val)
                        if dev == max_dev and cnt > best_cnt:
                            best_pair = (cn_val, intl_val)
                            best_cnt = cnt
                    if best_pair:
                        # 优先找同等级偏差中涉及雨/雪的误判对，更有意义
                        rain_pair = None
                        rain_cnt = 0
                        for (cn2, intl2), cnt2 in mismatch:
                            if cnt2 <= rain_cnt: continue
                            d2, _ = calc_weather_deviation(cn2, intl2)
                            if d2 == max_dev:
                                ta = WTH_TEXTS.get(cn2, {})
                                tb = WTH_TEXTS.get(intl2, {})
                                if ta.get('cat') in ('雨','雪') or tb.get('cat') in ('雨','雪'):
                                    rain_pair = (cn2, intl2)
                                    rain_cnt = cnt2
                        if rain_pair:
                            max_display = f'国内{rain_pair[0]}→国外{rain_pair[1]}({_cnt(rain_cnt)})'
                        else:
                            max_display = f'国内{best_pair[0]}→国外{best_pair[1]}({_cnt(best_cnt)})'
                    else:
                        max_display = f'偏差{int(max_dev)}级 {_cnt(s["dev_counts"][max_dev])}'
                else:
                    max_display = s['maxdiff']
            else:
                avg_display = round(s['sumdiff'] / s['n'], 2) if s['n'] else ''
                max_display = s['maxdiff']
            # 数值字段加单位
            if '天气现象' not in field and isinstance(avg_display, (int, float)):
                u = ''
                if '温度' in field or '体感' in field: u = '℃'
                elif '湿度' in field: u = '%'
                elif '风速' in field: u = 'm/s'
                elif '气压' in field: u = 'hPa'
                elif '降水' in field: u = 'mm'
                avg_display = f'{avg_display}{u}'
                if isinstance(max_display, (int, float)):
                    max_display = f'{max_display}{u}'

            ws2.append([field, m, period, s['total'], s['miss'], s['n'], s['ok'], rate, avg_display, max_display, s['maxcity']])
    for col, w in zip('ABCDEFGHIJK', [16, 10, 14, 8, 14, 10, 8, 10, 10, 10, 14]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = 'A2'

    # ---------- Sheet 前五偏差城市 ----------
    ws_top = wb.create_sheet('前五偏差城市')
    H3 = ['模块', '字段', '时效', '排名', '城市', 'CN→INTL', '偏差']
    ws_top.append(H3)
    for c in range(1, len(H3) + 1):
        ws_top.cell(row=1, column=c).font = Font(bold=True)
    PERIOD_ORDER_T = {p[0]: i for i, p in enumerate(PERIODS_24H)}
    for module in MODULES.keys():
        items = sorted(stat.items(), key=lambda x: (
            list(MODULES.keys()).index(x[0][1]) if x[0][1] in MODULES else 99,
            x[0][0],
            PERIOD_ORDER_T.get(x[0][2], 99)
        ))
        for (field, m, period), s in items:
            if m != module: continue
            top_items = list(s['top'].items())
            top_sorted = sorted(top_items, key=lambda x: (
                -abs(x[1][0]),
                -_weather_level_diff(x[1][1], x[1][2]),
                -CITY_RANK.get(x[0], 9999)
            ))[:5]

            for rank, (city, val) in enumerate(top_sorted, 1):
                d, cn_val, intl_val = val
                if cn_val and intl_val and '天气现象' in field:
                    pair_str = f'国内{cn_val}→国外{intl_val}'
                else:
                    pair_str = ''
                u = ''
                if '天气现象' not in field:
                    if '温度' in field or '体感' in field: u = '℃'
                    elif '湿度' in field: u = '%'
                    elif '风速' in field: u = 'm/s'
                    elif '气压' in field: u = 'hPa'
                    elif '降水' in field: u = 'mm'
                d_val = round(d, 2)
                ws_top.append([m, field, period or '', rank, city, pair_str, f'{d_val}{u}' if u else d_val])
    for col, w in zip('ABCDEFG', [10, 16, 14, 6, 14, 16, 10]):
        ws_top.column_dimensions[col].width = w
    ws_top.freeze_panes = 'A2'

    # ---------- Sheet3 说明 ----------
    ws3 = wb.create_sheet('说明')
    notes = [f'一致性比对报告 — {title}', '', f'配置文件: compare_config.yaml', '']
    notes.append(f'报告生成: {len(set(p[0] for p in allP))}个城市, {len(allP)}个数据点')
    notes.append(f'比对口径: {title}')
    notes.append('')
    if '阈值' in title:
        notes.append('一、一致判定阈值(来自配置):')
        for k, v in THRESHOLDS.items():
            notes.append(f'  {k} |差|≤{v}')
        notes.append('  天气现象 按语义映射比对(详见天气映射说明)')
        notes.append('  缺数据 标"缺数据",不计入一致率分母')
        notes.append('')
    notes.append('二、风速换算(来自配置):')
    notes.append(f"  enabled={WIND_CFG.get('enabled')}, divisor={WIND_CFG.get('divisor')} (海外km/h÷3.6→m/s)")
    notes.append('')
    notes.append('三、24小时时效分段: 短时效(1-6h) / 中时效(7-12h) / 长时效(13-24h)')
    notes.append('')
    notes.append('四、时次对齐: 24h按predict_time, 15天按predict_date')
    notes.append('')
    notes.append('五、天气现象语义映射比对规则（五分制评分）：')
    notes.append('  将国内外中文天气文字统一映射到(大类, 量级, 是否高影响)')
    notes.append('  评分→偏差规则:')
    notes.append('    5分 主天气一致+量级一致                     → 完全匹配')
    notes.append('    4分 主天气一致+量级差1级                     → 轻微量级偏差')
    notes.append('    3分 主天气不一致(晴↔多云等)，均非高影响       → 主天气不一致')
    notes.append('    2分 主天气错判(量级差≥2)，或降水vs非降水     → 明显偏差')
    notes.append('    1分 涉及高影响天气+同大类                   → 高影响偏差')
    notes.append('    0分 高影响天气+不同大类                     → 高影响漏报/错判')
    notes.append('  一致判定: 仅5分(偏差=0)算"一致"，其余全算"不一致"')
    notes.append('  高影响天气: 大雨/暴雨/大暴雨/特大暴雨/大雪/暴雪/雷暴/冰雹')
    notes.append('  完整映射对照表请见 compare_config.yaml → weather_mapping')
    notes.append('')
    notes.append('注: 修改 compare_config.yaml 后重跑即可更新阈值口径')
    for n in notes:
        ws3.append([n])
    if extra_notes:
        ws3.append([''])
        for n in extra_notes:
            ws3.append([n])
    ws3['A1'].font = Font(bold=True, size=13)

    wb.save(xlsx_path)
    return len(allP)

# =========================================================
# 主流程
# =========================================================

def load_cities():
    """读城市列表（去重），并设置 CITY_RANK（CSV 顺序，偏差相同时偏远地区优先）。
    返回 [(name, lon, lat), ...]。供一次性脚本和定时脚本复用"""
    global CITY_RANK
    cities = []
    seen = set()
    with open(CITY_CSV, 'r', encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            lon = r['Flon'].strip(); lat = r['Flat'].strip()
            key = (lon, lat)
            if key in seen: continue
            seen.add(key)
            cities.append((r['Fcityname_cn'].strip(), lon, lat))
    # 城市序号：CSV 顺序（北京、天津、上海...漠河、阿勒泰），偏差相同时偏远地区优先
    CITY_RANK = {name: i for i, (name, *_rest) in enumerate(cities)}
    return cities

def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    # 1. 读城市列表（去重）+ 设置偏远优先序号
    cities = load_cities()
    print(f"共 {len(cities)} 个城市, 开始请求实时 API...")

    # 2. 逐个城市拉数据 + 比对
    allP_strict = []
    allP_threshold = []
    ok_count = 0

    for idx, (name, lon, lat) in enumerate(cities, 1):
        cn_data, in_data = fetch_city(name, lon, lat)

        if cn_data is None:
            print(f"  [{idx}/{len(cities)}] ⏭️ {name} 国内接口失败, 跳过")
            continue
        if in_data is None:
            print(f"  [{idx}/{len(cities)}] ⏭️ {name} 国际接口失败, 跳过")
            continue

        # 同时构建两种口径的比对数据
        allP_strict += build_points(name, cn_data, in_data, strict=True)
        allP_threshold += build_points(name, cn_data, in_data, strict=False)
        ok_count += 1

        if idx % 10 == 0 or idx == len(cities):
            print(f"  [{idx}/{len(cities)}] {name} ✅")

    print(f"\n请求完成: 成功 {ok_count} 个城市, 跳过 {len(cities) - ok_count} 个")

    if not allP_strict:
        print("❌ 无有效数据, 不生成报告")
        return

    # 3. 生成严格相等报告
    n1 = gen_xlsx(allP_strict, '严格相等', XLSX_STRICT)
    print(f"\n✅ 严格相等报告: {XLSX_STRICT}")
    print(f"   数据点: {n1}")

    # 4. 生成阈值口径报告
    n2 = gen_xlsx(allP_threshold, '阈值口径', XLSX_THRESHOLD)
    print(f"✅ 阈值口径报告: {XLSX_THRESHOLD}")
    print(f"   数据点: {n2}")

    # 5. 导出数据明细CSV（替代原始JSON作为证据，长格式）
    with open(CSV_PATH, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(['城市', '模块', '字段', '时次', '国内值', '海外值', '差异', '是否一致', '备注', '时效分段'])
        for p in allP_threshold:
            w.writerow(p)
    print(f"✅ 数据明细CSV: {CSV_PATH}")
    print(f"   数据点: {len(allP_threshold)}")

    # 6. 快速打印阈值口径一致率
    print(f"\n{'='*60}")
    print(f"阈值口径 — 各字段一致率速览")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(XLSX_THRESHOLD)
    ws = wb['总结']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and not row[2]:  # 无时效分段的汇总行（实况/15天/AQI）
            print(f"  {str(row[0]):14s} {str(row[1]):8s}  一致率: {str(row[7]):>7s}  平均偏差: {str(row[8]):>8s}")

if __name__ == '__main__':
    main()
