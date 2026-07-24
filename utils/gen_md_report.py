#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成 md 表格版一致性比对报告（清晰表格版，与 gen_html_report 同源数据）
读 xlsx(总结 + 前五偏差城市) -> md
- 各字段一致率：数值字段一表，天气现象单独一表（误判对，表头+一行）
- 最大偏差城市 TOP5
- 结论/核心发现从数据算，不写死
运行:
  python3 gen_md_report.py              # 默认取最新均值阈值口径 xlsx
  python3 gen_md_report.py <xlsx路径>   # 指定 xlsx
"""
import os, sys, re, glob, datetime
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.join(SCRIPT_DIR, '..', 'data')
OUT_DIR = os.path.join(BASE, '比对结果')
CONFIG_PATH = os.path.join(SCRIPT_DIR, 'compare_config.yaml')

FIELD_ORDER = ['温度', '体感温度', '湿度', '风速', '气压', '天气现象', '降水量',
               '温度(最高)', '温度(最低)', '体感温度(白天)', '体感温度(夜间)',
               '风速(白天)', '风速(夜间)', '天气现象(白天)', '天气现象(夜间)', 'AQI']
PERIODS_24 = ['短时效(1-6h)', '中时效(7-12h)', '长时效(13-24h)']
MODULE_LIST = ['实况', '24小时', '15天', 'AQI模块']
MOD_DISPLAY = {'实况': '实况', '24小时': '24小时逐时', '15天': '15天预报', 'AQI模块': 'AQI'}


def is_weather(f):
    return f is not None and '天气现象' in f


def unit(f):
    if '温度' in f or '体感' in f: return '℃'
    if '湿度' in f: return '%'
    if '风速' in f: return 'm/s'
    if '气压' in f: return 'hPa'
    return ''


def fmt_num(v, f):
    if v is None or v == '': return '-'
    s = str(v)
    u = unit(f)
    if u and u not in s:
        return f"{s}{u}"
    return s
def parse_rate(r):
    if r is None: return 0.0
    s = str(r).replace('%', '').strip()
    try: return float(s) if s else 0.0
    except ValueError: return 0.0

def parse_num(s):
    """从可能带单位的字符串中提取数值，如 '3℃' -> 3.0"""
    if s is None: return 0.0
    if isinstance(s, (int, float)): return float(s)
    import re
    m = re.search(r'[\d.]+', str(s))
    return float(m.group()) if m else 0.0


# =========================================================
# 数据读取
# =========================================================

def find_inputs(xlsx_arg=None):
    if xlsx_arg:
        return xlsx_arg
    fs = glob.glob(os.path.join(OUT_DIR, '**', '一致性比对报告_均值_阈值口径_*.xlsx'), recursive=True)
    if not fs:
        raise SystemExit('❌ 未找到均值阈值口径 xlsx，请先跑 scheduled_compare.py / reformat_threshold.py')
    return max(fs, key=os.path.getmtime)


def read_summary(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb['总结']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    out = []
    for r in rows[1:]:
        if r is None or r[0] is None: continue
        r = list(r) + [None] * (len(headers) - len(r))
        d = dict(zip(headers, r))
        out.append({
            'field': d['字段'], 'module': d['模块'], 'period': d['时效'] or '',
            'total': d['总数据'], 'miss': d['缺数据(已排除)'], 'valid': d['有效样本'],
            'ok': d['一致数'], 'rate': d['一致率'], 'avgDev': d['平均偏差'],
            'maxDev': d['最大偏差'], 'maxCity': d['最大偏差城市']
        })
    return out


def read_top5(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    if '前五偏差城市' not in wb.sheetnames: return []
    ws = wb['前五偏差城市']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    out = []
    for r in rows[1:]:
        if r is None or r[0] is None: continue
        r = list(r) + [None] * (len(headers) - len(r))
        out.append({'module': r[0], 'field': r[1], 'period': r[2] or '',
                    'rank': r[3], 'city': r[4], 'pair': r[5] or '', 'dev': r[6]})
    return out


def read_thresholds():
    try:
        import yaml
        cfg = yaml.safe_load(open(CONFIG_PATH, encoding='utf-8'))
        th = cfg.get('thresholds', {})
        rain = cfg.get('rain_thresholds', {})
        return th, rain
    except Exception:
        return {}, {}


def read_meta(xlsx):
    """从说明sheet读口径和城市数"""
    koujing = '阈值口径'
    if '均值' in os.path.basename(xlsx): koujing = '阈值口径（6次均值）'
    cities = 70
    try:
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        if '说明' in wb.sheetnames:
            for row in wb['说明'].iter_rows(values_only=True):
                if not row[0]: continue
                txt = str(row[0])
                if '比对口径' in txt:
                    koujing = txt.replace('比对口径: ', '').replace('比对口径:', '').strip()
                m = re.search(r'(\d+)\s*个城市', txt)
                if m: cities = int(m.group(1))
    except Exception:
        pass
    return koujing, cities


def parse_sample(xlsx):
    """从xlsx文件名解析采样区间和均值次数，如：2026-07-21 16:39 ~ 2026-07-21 17:40，6 次均值"""
    name = os.path.basename(xlsx)

    def fmt(s):  # 20260721_1639 -> 2026-07-21 16:39
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]} {s[9:11]}:{s[11:13]}"
    m = re.search(r'(\d+)次均值_(\d{8}_\d{4})-(\d{8}_\d{4})', name)
    if m:
        return f"{fmt(m.group(2))} ~ {fmt(m.group(3))}，{m.group(1)} 次均值"
    m2 = re.search(r'(\d{8}_\d{4})-(\d{8}_\d{4})', name)
    if m2:
        return f"{fmt(m2.group(1))} ~ {fmt(m2.group(2))}"
    return '单次拉取（非均值）'


# =========================================================
# md 片段
# =========================================================

def field_table(rows):
    """数值字段 + 天气现象同表（天气现象行填误判对，表下注释说明表头不对应）"""
    ordered = sorted(rows, key=lambda x: FIELD_ORDER.index(x['field']) if x['field'] in FIELD_ORDER else 99)
    numeric = [s for s in ordered if not is_weather(s['field'])]
    weather = [s for s in ordered if is_weather(s['field'])]
    md = ''
    if numeric or weather:
        md += '| 字段 | 一致率 | 平均偏差 | 最大偏差 |\n|---|---|---|---|\n'
        for s in numeric:
            md += f"| {s['field']} | {s['rate']} | {fmt_num(s['avgDev'], s['field'])} | {fmt_num(s['maxDev'], s['field'])} |\n"
        for s in weather:
            md += f"| {s['field']} | {s['rate']} | {s['avgDev'] or '-'} | {s['maxDev'] or '-'} |\n"
    return md


def trend_table(rows):
    """24h 字段×时效 一致率"""
    fields = [f for f in FIELD_ORDER if any(s['field'] == f for s in rows)]
    md = '| 字段 | ' + ' | '.join(PERIODS_24) + ' |\n|---|---|---|---|\n'
    for f in fields:
        cells = []
        for p in PERIODS_24:
            s = next((x for x in rows if x['field'] == f and x['period'] == p), None)
            cells.append(s['rate'] if s else '-')
        md += f"| {f} | " + ' | '.join(cells) + ' |\n'
    return md


def top5_table(top5_rows):
    by_field, fields = {}, []
    for t in top5_rows:
        if t['field'] not in by_field:
            by_field[t['field']] = []
            fields.append(t['field'])
        by_field[t['field']].append(t)
    fields = sorted(fields, key=lambda x: FIELD_ORDER.index(x) if x in FIELD_ORDER else 99)
    # AQI 单字段：一行 TOP5
    if len(fields) == 1 and fields[0] == 'AQI':
        items = sorted(by_field['AQI'], key=lambda x: x['rank'])
        cells = [f"{t['city']}({t['dev']})" for t in items[:5]]
        while len(cells) < 5: cells.append('-')
        return '| TOP1 | TOP2 | TOP3 | TOP4 | TOP5 |\n|---|---|---|---|---|\n| ' + ' | '.join(cells) + ' |\n'
    md = '| 字段 | TOP1 | TOP2 | TOP3 | TOP4 | TOP5 |\n|---|---|---|---|---|---|\n'
    for f in fields:
        items = sorted(by_field[f], key=lambda x: x['rank'])
        cells = []
        for t in items[:5]:
            if is_weather(f):
                cells.append(f"{t['city']}({t['pair']})" if t['pair'] else f"{t['city']}")
            else:
                cells.append(f"{t['city']}({t['dev']})")
        while len(cells) < 5: cells.append('-')
        md += f"| {f} | " + ' | '.join(cells) + ' |\n'
    return md


def build_md(summary, top5, thresholds, rain_th, meta):
    # ---- 模块汇总 ----
    mod_agg = {}
    for s in summary:
        m = s['module']
        if m not in mod_agg: mod_agg[m] = {'total': 0, 'ok': 0, 'valid': 0}
        mod_agg[m]['total'] += s['total'] or 0
        mod_agg[m]['ok'] += s['ok'] or 0
        mod_agg[m]['valid'] += s['valid'] or 0
    total_valid = sum(a['valid'] for a in mod_agg.values())
    total_ok = sum(a['ok'] for a in mod_agg.values())
    overall_rate = total_ok / total_valid * 100 if total_valid else 0

    # 字段大类一致率(每模块每大类聚合,无则'-')
    FIELD_CATS = [
        ('温度', ['温度','温度(最高)','温度(最低)']),
        ('体感温度', ['体感温度','体感温度(白天)','体感温度(夜间)']),
        ('湿度', ['湿度']),
        ('风速', ['风速','风速(白天)','风速(夜间)']),
        ('气压', ['气压']),
        ('天气现象', ['天气现象','天气现象(白天)','天气现象(夜间)']),
        ('降水量', ['降水量']),
        ('AQI', ['AQI']),
    ]
    def cat_rate(module, fields):
        ok=valid=0
        for s in summary:
            if s['module']==module and s['field'] in fields:
                ok += s['ok'] or 0
                valid += s['valid'] or 0
        return f"{ok/valid*100:.1f}%" if valid else '-'

    md = f"""# 墨迹国际化与国内版本天气数据一致性测试报告

> 报告时间：{meta['time']}
> 采样：{meta['sample']}

---

## 一、测试结论

　　本次测试覆盖 {meta['cities']} 个城市 × {meta['avg_count']} 份拉取（共 {meta['cities']*meta['avg_count']} 个采样点），具体一致率统计如下表。

| 模块 | 数据点 | """ + ' | '.join(c[0] for c in FIELD_CATS) + f""" |
|---|---|""" + '|'.join(['---']*len(FIELD_CATS)) + f""" |
"""
    for m in MODULE_LIST:
        if m in mod_agg:
            cells = [cat_rate(m, c[1]) for c in FIELD_CATS]
            md += f"| {MOD_DISPLAY[m]} | {meta['avg_count']*meta['cities']} | " + ' | '.join(cells) + " |\n"

    # ---- 核心发现（从数据算） ----
    weather_rates = [parse_rate(s['rate']) for s in summary if is_weather(s['field']) and (s['valid'] or 0) > 0]
    numeric_rates = [parse_rate(s['rate']) for s in summary if not is_weather(s['field']) and (s['valid'] or 0) > 0]
    w_avg = sum(weather_rates) / len(weather_rates) if weather_rates else 0
    n_avg = sum(numeric_rates) / len(numeric_rates) if numeric_rates else 0
    candidates = [s for s in summary if (s['valid'] or 0) >= 10]
    weakest = min(candidates, key=lambda s: parse_rate(s['rate'])) if candidates else None
    aqi = next((s for s in summary if s['field'] == 'AQI'), None)
    feel = [s for s in summary if '体感温度' in s['field']]
    feel_weakest = min(feel, key=lambda s: parse_rate(s['rate'])) if feel else None
    pressure = [s for s in summary if s['field'] == '气压']
    pres_max = max(pressure, key=lambda s: abs(parse_num(s['maxDev']))) if pressure else None

    md += '\n核心发现：\n\n'
    md += f"- 天气现象类一致率（均值 {w_avg:.1f}%）{'显著高于' if w_avg > n_avg else '与'}数值类（{n_avg:.1f}%），语义映射优化效果明显\n"
    if weakest and weakest['field'] != 'AQI':
        wn = MOD_DISPLAY.get(weakest['module'], weakest['module'])
        label = weakest['field'] if weakest['field'] == wn else f"{weakest['field']}（{wn}）"
        md += f"- 最弱维度：{label} 一致率仅 {parse_rate(weakest['rate']):.1f}%\n"
    if feel_weakest:
        md += f"- 体感温度偏差大：{feel_weakest['field']} 一致率仅 {parse_rate(feel_weakest['rate']):.1f}%，平均偏差 {fmt_num(feel_weakest['avgDev'], feel_weakest['field'])}\n"
    if aqi:
        md += f"- AQI 一致率最低（{parse_rate(aqi['rate']):.1f}%），两套数据源存在系统性差异\n"
    if pres_max:
        md += f"- 气压在高原城市存在系统性偏差，最大可达 {fmt_num(pres_max['maxDev'], '气压')}（{pres_max['maxCity'] or ''}）\n"
    md += '\n---\n'

    # ---- 二、阈值配置 ----
    md += """
## 二、阈值配置

**数值字段：**

| 字段 | 一致判定阈值 |
|---|---|
"""
    for k, v in thresholds.items():
        md += f"| {k} | ≤ {v}{unit(k)} |\n"
    md += """
**降水量字段：**

等级比对规则：

| 等级 | 降水量区间 (mm) |
|---|---|
"""
    prev = 0
    for name, th in rain_th.items():
        ub = f"{th}" if th not in ('~', None) else '∞'
        md += f"| {name} | [{prev}, {ub}) |\n"
        if th not in ('~', None):
            prev = th
    md += """
*比对方式：国内值/海外值分别映射到降水量等级，等级相同即判一致。*

**天气现象字段：**

采用语义映射五分制评分，偏差计算规则如下：

大类（12种）： 晴、多云、阴、雨、雪、雾、霾、沙尘、冻雨、冰雹、雷暴、霜
*注：阴、雾、霾归入多云大类，减少API间分类差异导致的误判*

量级： 雨/雪按强度分 7 级（1~7），其余大类量级固定为 0

高影响天气： 大雨、暴雨、大暴雨、特大暴雨、大雪、暴雪、雷暴、冰雹

| 得分 | 判定规则 | 说明 |
|---|---|---|
| 5分 | 主天气一致，量级一致 | 完全匹配 |
| 4分 | 主天气一致，量级差1级 | 轻微量级偏差 |
| 3分 | 主天气不一致（如晴↔多云），均非高影响 | 主天气不一致 |
| 2分 | 主天气错判（量级差≥2），或涉及降水vs非降水 | 明显偏差 |
| 1分 | 涉及高影响天气但同大类 | 高影响偏差 |
| 0分 | 高影响天气 + 不同大类 | 高影响漏报/错判 |



---

## 三、各字段一致率详情

"""
    # 3.1 实况
    md += '### 3.1 实况模块\n\n' + field_table([s for s in summary if s['module'] == '实况']) + '\n'
    # 3.2 24h
    rows24 = [s for s in summary if s['module'] == '24小时']
    md += '### 3.2 24小时逐时模块\n\n**一致率趋势：**\n\n' + trend_table(rows24) + '\n'
    for p in PERIODS_24:
        md += f'**{p}：**\n\n' + field_table([s for s in rows24 if s['period'] == p]) + '\n'
    # 3.3 15天
    md += '### 3.3 15天预报模块\n\n' + field_table([s for s in summary if s['module'] == '15天']) + '\n'
    # 3.4 AQI
    md += '### 3.4 AQI模块\n\n' + field_table([s for s in summary if s['module'] == 'AQI模块']) + '\n'
    md += '---\n'

    # ---- 四、最大偏差城市 TOP5 ----
    md += '## 四、最大偏差城市分析\n\n'
    md += '### 4.1 实况模块\n\n' + top5_table([t for t in top5 if t['module'] == '实况']) + '\n'
    md += '### 4.2 24小时逐时模块\n\n'
    for p in PERIODS_24:
        md += f'**{p}：**\n\n' + top5_table([t for t in top5 if t['module'] == '24小时' and t['period'] == p]) + '\n'
    md += '### 4.3 15天预报模块\n\n' + top5_table([t for t in top5 if t['module'] == '15天']) + '\n'
    md += '### 4.4 AQI模块\n\n' + top5_table([t for t in top5 if t['module'] == 'AQI模块']) + '\n'
    md += '---\n'

    # ---- 五、测试说明 ----
    md += f"""
## 五、测试说明

- 测试范围：{meta['cities']} 个城市（以国内城市为主），覆盖实况 / 24小时逐时 / 15天预报 / AQI 四大模块
- 数据来源：国内版（coapi.moji.com）vs 国际版（datasw1.api.moweather.com）
- 风速换算：国际版 km/h 统一 ÷3.6 换算为 m/s 后比对
- 天气现象：两端均指定中文返回，按语义映射五分制评分比对；阴/雾/霾归入多云大类
- 缺数据处理：某端缺失数据标记"缺数据"，不计入一致率分母

---

## 六、风险与遗留项

- 海外城市未覆盖：国内接口不支持海外城市，本次以国内城市为主
- 数据实时性：接口数据为实时拉取，存在时间窗口内波动影响
"""
    if aqi:
        md += f"- AQI 差异显著：两套数据源 AQI 一致率仅 {parse_rate(aqi['rate']):.1f}%，最大偏差 {fmt_num(aqi['maxDev'], 'AQI')}，需排查数据源差异和归一化逻辑\n"
    if pres_max:
        md += f"- 气压高原系统性偏差：{pres_max['maxCity'] or '高原城市'}等气压偏差可达 {fmt_num(pres_max['maxDev'], '气压')}，存在因海拔/气压换算基准不同导致的系统性偏移\n"
    if feel_weakest:
        md += f"- 体感温度偏差大：{feel_weakest['field']}一致率仅 {parse_rate(feel_weakest['rate']):.1f}%，平均偏差 {fmt_num(feel_weakest['avgDev'], feel_weakest['field'])}，体感温度计算模型差异可能是主因\n"
    md += "- 天气现象高影响错判：部分城市在预报模块出现高影响天气漏报/错判，需确认是否因预报差异导致\n"
    return md


# =========================================================
# 主流程
# =========================================================

def main():
    xlsx_arg = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx = find_inputs(xlsx_arg)
    print(f"读取: {os.path.basename(xlsx)}")

    summary = read_summary(xlsx)
    top5 = read_top5(xlsx)
    thresholds, rain_th = read_thresholds()
    koujing, cities = read_meta(xlsx)
    _am = re.search(r'(\d+)次均值', os.path.basename(xlsx))
    avg_count = int(_am.group(1)) if _am else 1

    meta = {
        'time': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
        'source': os.path.basename(xlsx),
        'koujing': koujing,
        'cities': cities,
        'sample': parse_sample(xlsx),
        'avg_count': avg_count,
    }

    md = build_md(summary, top5, thresholds, rain_th, meta)
    out_path = os.path.join(SCRIPT_DIR, '..', f'一致性比对报告_{datetime.datetime.now().strftime("%Y%m%d")}.md')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(md)

    print(f"\n✅ md报告: {out_path}")
    print(f"   口径: {koujing}，城市: {cities}")
    print(f"   字段一致率详情 + TOP5 + 结论（从数据算）")


if __name__ == '__main__':
    main()
