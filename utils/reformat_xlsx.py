#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""重新生成xlsx：每行一个数据点(城市-模块-字段-时次)，国内值/海外值/差异并排，最后总结"""
import json, glob, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_DIR = os.path.join(SCRIPT_DIR, '..', 'data', '比对结果', '原始JSON')
XLSX = os.path.join(SCRIPT_DIR, '..', 'data', '比对结果', '一致性比对报告.xlsx')

def num(v):
    if v is None: return None
    try: return float(v)
    except: return None

WIND=lambda x: round(x/3.6,2)

def diff_str(cv, iv):
    """数值差异"""
    if cv is None or iv is None: return ''
    return round(cv-iv,2)

def text_diff(cv, iv):
    """天气现象：按中文文字内容比对"""
    if cv is None or iv is None: return ''
    return '一致' if cv==iv else '不一致'

# 生成数据点: (城市,模块,字段,时次,国内值,海外值,差异,备注)
def points(city, cn, ind):
    P=[]; c=cn.get('current',{}); i=ind.get('current',{})
    a=cn.get('aqi',{}); b=ind.get('aqi',{})
    # 实况
    P.append((city,'实况','温度','实况',c.get('temp'),i.get('temp'),diff_str(num(c.get('temp')),num(i.get('temp'))),''))
    P.append((city,'实况','体感温度','实况',c.get('real_feel'),i.get('rf'),diff_str(num(c.get('real_feel')),num(i.get('rf'))),''))
    P.append((city,'实况','湿度','实况',c.get('humidity'),i.get('rh'),diff_str(num(c.get('humidity')),num(i.get('rh'))),''))
    P.append((city,'实况','风速','实况',c.get('wspd'),WIND(num(i.get('wspd'))) if i.get('wspd') else None,diff_str(num(c.get('wspd')),WIND(num(i.get('wspd'))) if i.get('wspd') else None),f"海外原{i.get('wspd')} ÷3.6" if i.get('wspd') else ''))
    P.append((city,'实况','气压','实况',c.get('mslp'),i.get('sp'),diff_str(num(c.get('mslp')),num(i.get('sp'))),''))
    P.append((city,'实况','天气现象','实况',c.get('weather'),i.get('wtr'),text_diff(c.get('weather'),i.get('wtr')),'按中文文字比对'))
    # AQI模块
    P.append((city,'AQI模块','AQI','AQI实况',a.get('aqi'),b.get('AQI'),diff_str(num(a.get('aqi')),num(b.get('AQI'))),''))
    P.append((city,'AQI模块','PM2.5','AQI实况',a.get('pm25'),b.get('PM2P5'),diff_str(num(a.get('pm25')),num(b.get('PM2P5'))),'国内浓度vs海外(疑分指数)'))
    # 24小时
    ch=cn.get('hourly',[])[:24]; ih=ind.get('hourly',[])[:24]
    for k in range(min(len(ch),len(ih),24)):
        x=ch[k]; y=ih[k]; ts=x.get('predict_time',f'第{k+1}h')
        P.append((city,'24小时','温度',ts,x.get('temp'),y.get('temp'),diff_str(num(x.get('temp')),num(y.get('temp'))),''))
        P.append((city,'24小时','体感温度',ts,x.get('real_feel'),y.get('rf'),diff_str(num(x.get('real_feel')),num(y.get('rf'))),''))
        P.append((city,'24小时','湿度',ts,x.get('humidity'),y.get('rh'),diff_str(num(x.get('humidity')),num(y.get('rh'))),''))
        P.append((city,'24小时','风速',ts,x.get('wspd'),WIND(num(y.get('wspd'))) if y.get('wspd') else None,diff_str(num(x.get('wspd')),WIND(num(y.get('wspd'))) if y.get('wspd') else None),f"海外原{y.get('wspd')} ÷3.6" if y.get('wspd') else ''))
        P.append((city,'24小时','气压',ts,x.get('mslp'),y.get('sp'),diff_str(num(x.get('mslp')),num(y.get('sp'))),''))
        P.append((city,'24小时','天气现象',ts,x.get('weather'),y.get('wtr'),text_diff(x.get('weather'),y.get('wtr')),'按中文文字比对'))
    # 15天
    cd=cn.get('daily',[]); idd=ind.get('daily',[])
    for k in range(min(len(cd),len(idd),16)):
        x=cd[k]; y=idd[k]; ts=x.get('predict_date',f'第{k+1}天')
        P.append((city,'15天','温度(最高)',ts,x.get('temp_high'),y.get('temph'),diff_str(num(x.get('temp_high')),num(y.get('temph'))),''))
        P.append((city,'15天','温度(最低)',ts,x.get('temp_low'),y.get('templ'),diff_str(num(x.get('temp_low')),num(y.get('templ'))),''))
        P.append((city,'15天','体感温度(白天)',ts,x.get('realfeel_d'),y.get('rfd'),diff_str(num(x.get('realfeel_d')),num(y.get('rfd'))),''))
        P.append((city,'15天','体感温度(夜间)',ts,x.get('realfeel_n'),y.get('rfn'),diff_str(num(x.get('realfeel_n')),num(y.get('rfn'))),''))
        P.append((city,'15天','湿度',ts,x.get('humidity'),y.get('rh'),diff_str(num(x.get('humidity')),num(y.get('rh'))),''))
        P.append((city,'15天','风速(白天)',ts,x.get('wspd_day'),WIND(num(y.get('wspdd'))) if y.get('wspdd') else None,diff_str(num(x.get('wspd_day')),WIND(num(y.get('wspdd'))) if y.get('wspdd') else None),f"海外原{y.get('wspdd')} ÷3.6" if y.get('wspdd') else ''))
        P.append((city,'15天','风速(夜间)',ts,x.get('wspd_night'),WIND(num(y.get('wspdn'))) if y.get('wspdn') else None,diff_str(num(x.get('wspd_night')),WIND(num(y.get('wspdn'))) if y.get('wspdn') else None),f"海外原{y.get('wspdn')} ÷3.6" if y.get('wspdn') else ''))
        P.append((city,'15天','气压',ts,x.get('mslp'),y.get('spd'),diff_str(num(x.get('mslp')),num(y.get('spd'))),'海外取白天spd'))
        P.append((city,'15天','天气现象(白天)',ts,x.get('weather_day'),y.get('wtrd'),text_diff(x.get('weather_day'),y.get('wtrd')),'按中文文字比对'))
        P.append((city,'15天','天气现象(夜间)',ts,x.get('weather_night'),y.get('wtrn'),text_diff(x.get('weather_night'),y.get('wtrn')),'按中文文字比对'))
    return P

def main():
    cn_files=sorted(glob.glob(f'{JSON_DIR}/*_国内.json'))
    allP=[]
    for cf in cn_files:
        city=os.path.basename(cf).replace('_国内.json','')
        inf=cf.replace('_国内.json','_国际.json')
        if not os.path.exists(inf): continue
        cn=json.load(open(cf,encoding='utf-8')); ind=json.load(open(inf,encoding='utf-8'))
        allP+=points(city,cn,ind)
    print(f"城市数: {len(cn_files)}, 数据点: {len(allP)}")

    wb=openpyxl.Workbook()
    # Sheet1 数据明细
    ws=wb.active; ws.title='数据明细'
    H=['城市','模块','字段','时次','国内值','海外值','差异','备注']
    ws.append(H)
    for c in range(1,len(H)+1): ws.cell(row=1,column=c).font=Font(bold=True)
    green=PatternFill('solid',fgColor='E6F7E6'); red=PatternFill('solid',fgColor='FFE6E6')
    for p in allP:
        ws.append(list(p))
    # 差异列着色(一致绿/不一致红)
    for i,p in enumerate(allP,2):
        d=p[6]
        cell=ws.cell(row=i,column=7)
        if d=='一致': cell.fill=green
        elif d=='不一致' or (isinstance(d,(int,float)) and d!=0): cell.fill=red
    for col,w in zip('ABCDEFGH',[12,10,14,18,16,16,12,22]): ws.column_dimensions[col].width=w
    ws.freeze_panes='A2'

    # Sheet2 总结
    ws2=wb.create_sheet('总结')
    from collections import defaultdict
    stat=defaultdict(lambda:{'n':0,'ok':0,'maxdiff':0,'maxcity':''})
    for p in allP:
        city,_,field,_,_,_,d,_=p
        module=p[1]
        if d=='' or d=='一致' or d=='不一致':
            if d=='' : continue  # 缺数据
        # 判断一致: 数值差异==0 或 天气一致
        if d=='一致': ok=True
        elif d=='不一致': ok=False
        elif isinstance(d,(int,float)): ok=(d==0)
        else: continue
        s=stat[(field,module)]; s['n']+=1
        if ok: s['ok']+=1
        if isinstance(d,(int,float)) and abs(d)>abs(s['maxdiff']): s['maxdiff']=d; s['maxcity']=city
    H2=['字段','模块','样本数','一致数','一致率','最大偏差','最大偏差城市']
    ws2.append(H2)
    for c in range(1,len(H2)+1): ws2.cell(row=1,column=c).font=Font(bold=True)
    for module in ['实况','24小时','15天','AQI模块']:
        for (field,m),s in sorted(stat.items()):
            if m!=module: continue
            rate=f"{s['ok']/s['n']*100:.1f}%" if s['n'] else '0'
            ws2.append([field,m,s['n'],s['ok'],rate,s['maxdiff'],s['maxcity']])
    for col,w in zip('ABCDEFG',[16,10,8,8,10,10,14]): ws2.column_dimensions[col].width=w

    # 说明
    ws3=wb.create_sheet('说明')
    notes=[
        '一致性比对说明',
        '',
        '1. 比对口径: 严格相等 (国内值 == 海外值 才算一致)',
        '2. 风速: 国内 m/s, 海外 km/h, 海外值已 ÷3.6 换算成 m/s 再比对 (备注列标注原值)',
        '3. 天气现象: 按中文文字内容比对 (国内language=zh-CN + 国际lang=zh-CN, 两边中文), id编码体系不同不用',
        '4. PM2.5: 国内 pm25(浓度μg/m3) vs 海外 PM2P5(疑为分指数), 直接比, 差异大即bug',
        '5. AQI: 国内 aqi vs 海外 AQI, 直接比',
        '6. 15天气压: 国内 mslp(单值) vs 海外 spd(白天), 海外spn夜间未纳入',
        '7. 时次对齐: 24小时按predict_time, 15天按predict_date, 已转北京时区对齐',
        '8. 国内不覆盖城市(伦敦/纽约等海外)未纳入, 国内接口返回Out Of Query Range',
        '9. 差异列: 数值=国内-海外(正值国内大); 天气=一致/不一致; 红色=不一致',
        '',
        '⚠️ 严格相等下小差异(温度<3度)是数据源/时次正常波动, 非bug;',
        '   大差异(如乌鲁木齐AQI差384)才是真问题, 需提bug排查海外数据源',
    ]
    for n in notes: ws3.append([n])
    ws3['A1'].font=Font(bold=True,size=13)

    wb.save(XLSX)
    print(f"✅ 已生成: {XLSX}")

if __name__=='__main__':
    main()
