#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
国内/国际天气数据一致性批量比对
- 8字段 × 模块，严格相等
- 风速: 国际 wspd ÷3.6 换算 m/s
- PM2.5: 国内 pm25 vs 国际 PM2P5 直接比
- 天气现象: id 比 (weather_id vs wtrid), 文字保留
- 输出: Excel(汇总+明细) + 每城市原始JSON
- 明细附国际精简URL+国内URL可追溯
"""
import hashlib, hmac, subprocess, json, csv, os, time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ============== 配置 ==============
PASSWORD_CN='49ff9a4e5e8bd5e8ce9e057c5adc5d2d'
PASSWORD_IN='923ffbda8b65bf0f8e126824d050887a'
TOKEN_CN='cc920d85f8fbb762b6c705375add6c32'
TOKEN_IN='b88b7a5375e293671270016fe556a4b5'
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.join(SCRIPT_DIR, '..', 'data', '墨迹国际化与国内版本天气数据一致性测试')
CSV_PATH = os.path.join(BASE, '天气一致性测试城市_热门城市筛选.csv')
OUT_DIR = os.path.join(BASE, '比对结果')
JSON_DIR = os.path.join(OUT_DIR, '原始JSON')
XLSX_PATH = os.path.join(OUT_DIR, '一致性比对报告.xlsx')
# =================================

def cn_key(lat,lon): return hashlib.md5((PASSWORD_CN+'0'+lat+lon).encode()).hexdigest()
def in_key(lat,lon): return hmac.new(PASSWORD_IN.encode(),('0'+lat+lon).encode(),hashlib.sha256).hexdigest()

def cn_url(lat,lon):
    return f'http://coapi.moji.com/whapi/v2/weather?timestamp=0&token={TOKEN_CN}&lat={lat}&lon={lon}&key={cn_key(lat,lon)}'

def in_url_full(lat,lon):
    return f'https://datasw1.api.moweather.com/whapi/in/weather?token={TOKEN_IN}&lon={lon}&lat={lat}&lang=zh-CN&current=1&hourly=24&daily=15&aqi=1&metric=true&ts=0&key={in_key(lat,lon)}'

def in_url_module(lat,lon,module):
    p={'current':'current=1','hourly':'hourly=24','daily':'daily=15','aqi':'aqi=1'}[module]
    return f'https://datasw1.api.moweather.com/whapi/in/weather?token={TOKEN_IN}&lon={lon}&lat={lat}&lang=zh-CN&{p}&ts=0&key={in_key(lat,lon)}'

def fetch_full(url, retry=2):
    for _ in range(retry+1):
        r=subprocess.run(['curl','-sk','--max-time','25',url],capture_output=True,text=True)
        try:
            d=json.loads(r.stdout)
            if d.get('code')==0: return 0, d.get('msg',''), d.get('data',{})
            return d.get('code'), d.get('msg',''), None
        except: pass
        time.sleep(0.5)
    return -1,'request_fail',None

def to_num(v):
    if v is None or v=='' : return None
    try: return float(v)
    except: return None

WIND_CONV = lambda x: round(x/3.6, 2)

# 比对一条，返回明细行
def row(city, field, module, ts, cnv, inv, inurl, cnurl, conv=None):
    cv = to_num(cnv)
    iv = to_num(inv)
    if cv is None or iv is None:
        return [city,field,module,ts,cnv,inv,'' if not conv else '', '缺数据','', inurl, cnurl]
    iv_conv = conv(iv) if conv else iv
    same = (cv == iv_conv)
    diff = round(cv - iv_conv, 2)
    return [city,field,module,ts, cv, iv, iv_conv if conv else '',
            '一致' if same else '不一致', diff, inurl, cnurl]

def compare_city(name, lon, lat, cn, ind, cnurl):
    rows=[]
    # ========= 实况 current =========
    inu = in_url_module(lat,lon,'current')
    c=cn.get('current',{}); i=ind.get('current',{})
    rows.append(row(name,'温度','实况','实况',c.get('temp'),i.get('temp'),inu,cnurl))
    rows.append(row(name,'体感温度','实况','实况',c.get('real_feel'),i.get('rf'),inu,cnurl))
    rows.append(row(name,'湿度','实况','实况',c.get('humidity'),i.get('rh'),inu,cnurl))
    rows.append(row(name,'风速','实况','实况',c.get('wspd'),i.get('wspd'),inu,cnurl,WIND_CONV))
    rows.append(row(name,'气压','实况','实况',c.get('mslp'),i.get('sp'),inu,cnurl))
    rows.append(row(name,'天气现象','实况','实况',c.get('weather_id'),i.get('wtrid'),inu,cnurl))

    # ========= 24小时 hourly =========
    inu = in_url_module(lat,lon,'hourly')
    ch=cn.get('hourly',[])[:24]; ih=ind.get('hourly',[])[:24]
    for k in range(24):
        if k>=len(ch) or k>=len(ih): break
        a=ch[k]; b=ih[k]
        ts=a.get('predict_time',f'第{k+1}h')
        rows.append(row(name,'温度','24小时',ts,a.get('temp'),b.get('temp'),inu,cnurl))
        rows.append(row(name,'体感温度','24小时',ts,a.get('real_feel'),b.get('rf'),inu,cnurl))
        rows.append(row(name,'湿度','24小时',ts,a.get('humidity'),b.get('rh'),inu,cnurl))
        rows.append(row(name,'风速','24小时',ts,a.get('wspd'),b.get('wspd'),inu,cnurl,WIND_CONV))
        rows.append(row(name,'气压','24小时',ts,a.get('mslp'),b.get('sp'),inu,cnurl))
        rows.append(row(name,'天气现象','24小时',ts,a.get('weather_id'),b.get('wtrid'),inu,cnurl))

    # ========= 15天 daily =========
    inu = in_url_module(lat,lon,'daily')
    cd=cn.get('daily',[]); idd=ind.get('daily',[])
    n=min(len(cd),len(idd),16)
    for k in range(n):
        a=cd[k]; b=idd[k]
        ts=a.get('predict_date',f'第{k+1}天')
        rows.append(row(name,'温度(最高)','15天',ts,a.get('temp_high'),b.get('temph'),inu,cnurl))
        rows.append(row(name,'温度(最低)','15天',ts,a.get('temp_low'),b.get('templ'),inu,cnurl))
        rows.append(row(name,'体感温度(白天)','15天',ts,a.get('realfeel_d'),b.get('rfd'),inu,cnurl))
        rows.append(row(name,'体感温度(夜间)','15天',ts,a.get('realfeel_n'),b.get('rfn'),inu,cnurl))
        rows.append(row(name,'湿度','15天',ts,a.get('humidity'),b.get('rh'),inu,cnurl))
        rows.append(row(name,'风速(白天)','15天',ts,a.get('wspd_day'),b.get('wspdd'),inu,cnurl,WIND_CONV))
        rows.append(row(name,'风速(夜间)','15天',ts,a.get('wspd_night'),b.get('wspdn'),inu,cnurl,WIND_CONV))
        rows.append(row(name,'气压','15天',ts,a.get('mslp'),b.get('spd'),inu,cnurl))
        rows.append(row(name,'天气现象(白天)','15天',ts,a.get('weather_id_day'),b.get('wtridd'),inu,cnurl))
        rows.append(row(name,'天气现象(夜间)','15天',ts,a.get('weather_id_night'),b.get('wtridn'),inu,cnurl))

    # ========= AQI =========
    inu = in_url_module(lat,lon,'aqi')
    ca=cn.get('aqi',{}); ia=ind.get('aqi',{})
    rows.append(row(name,'AQI','AQI模块','AQI实况',ca.get('aqi'),ia.get('AQI'),inu,cnurl))
    rows.append(row(name,'PM2.5','AQI模块','AQI实况',ca.get('pm25'),ia.get('PM2P5'),inu,cnurl))
    return rows

def main():
    os.makedirs(JSON_DIR, exist_ok=True)
    # 读CSV 去重
    cities=[]
    seen=set()
    with open(CSV_PATH,'r',encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            lon=r['Flon'].strip(); lat=r['Flat'].strip()
            key=(lon,lat)
            if key in seen: continue
            seen.add(key)
            cities.append((r['Fcityname_cn'].strip(),lon,lat))
    print(f"共 {len(cities)} 个唯一城市")

    all_rows=[]
    fail=[]
    uncovered=[]
    for idx,(name,lon,lat) in enumerate(cities,1):
        cnurl=cn_url(lat,lon)
        cc,cm,cn=fetch_full(cnurl)
        ic,im,ind=fetch_full(in_url_full(lat,lon))
        if cc!=0:
            uncovered.append((name,cm))
            print(f"[{idx}/{len(cities)}] ⏭️ {name} 国内不覆盖({cm})")
            continue
        if ic!=0:
            fail.append((name,im)); print(f"[{idx}/{len(cities)}] ❌ {name} 国际失败({im})")
            continue
        # 存原始JSON
        with open(f'{JSON_DIR}/{name}_国内.json','w',encoding='utf-8') as f: json.dump(cn,f,ensure_ascii=False)
        with open(f'{JSON_DIR}/{name}_国际.json','w',encoding='utf-8') as f: json.dump(ind,f,ensure_ascii=False)
        all_rows += compare_city(name,lon,lat,cn,ind,cnurl)
        if idx%10==0: print(f"[{idx}/{len(cities)}] 已完成 {name}")
    print(f"\n国内不覆盖城市({len(uncovered)}): {[u[0] for u in uncovered]}")
    print(f"国际失败城市: {fail if fail else '无'}")
    print(f"明细总行数: {len(all_rows)}")

    # ===== 生成Excel =====
    wb=openpyxl.Workbook()
    # Sheet1 汇总
    ws1=wb.active; ws1.title='汇总'
    # 统计 by (字段,模块)
    from collections import defaultdict
    stat=defaultdict(lambda:{'n':0,'ok':0,'maxdiff':0,'maxcity':''})
    for r in all_rows:
        field,module=r[1],r[2]; same=r[7]; diff=r[8]
        if same=='缺数据': continue
        key=(field,module); s=stat[key]
        s['n']+=1
        if same=='一致': s['ok']+=1
        if isinstance(diff,(int,float)) and abs(diff)>abs(s['maxdiff']):
            s['maxdiff']=diff; s['maxcity']=r[0]
    headers1=['字段','模块','样本数','一致数','一致率','最大偏差','最大偏差城市']
    ws1.append(headers1)
    for col in range(1,len(headers1)+1):
        ws1.cell(row=1,column=col).font=Font(bold=True)
    order=['实况','24小时','15天','AQI模块']
    for module in order:
        for (field,m),s in sorted(stat.items()):
            if m!=module: continue
            rate=f"{s['ok']/s['n']*100:.1f}%" if s['n'] else '0'
            ws1.append([field,m,s['n'],s['ok'],rate,s['maxdiff'],s['maxcity']])
    # 列宽
    for col,w in zip('ABCDEFG',[16,10,8,8,8,10,14]): ws1.column_dimensions[col].width=w

    # Sheet2 明细
    ws2=wb.create_sheet('明细')
    headers2=['城市','字段','模块','时次','国内值','国际原值','换算值','是否一致','偏差','国际精简URL','国内URL']
    ws2.append(headers2)
    for col in range(1,len(headers2)+1):
        ws2.cell(row=1,column=col).font=Font(bold=True)
    for r in all_rows: ws2.append(r)
    # 一致标绿 不一致标红 缺数据灰
    green=PatternFill('solid',fgColor='E6F7E6'); red=PatternFill('solid',fgColor='FFE6E6'); gray=PatternFill('solid',fgColor='F0F0F0')
    for i,r in enumerate(all_rows,2):
        c=ws2[f'H{i}']
        if r[7]=='一致': c.fill=green
        elif r[7]=='不一致': c.fill=red
        else: c.fill=gray
    for col,w in zip('ABCDEFGHIJK',[10,14,8,16,10,10,10,9,8,30,30]): ws2.column_dimensions[col].width=w
    ws2.freeze_panes='A2'

    wb.save(XLSX_PATH)
    print(f"\n✅ 报告已生成: {XLSX_PATH}")
    print(f"   原始JSON: {JSON_DIR}/")

if __name__=='__main__':
    main()
